# cashflow_projection_report.py
import frappe
from frappe import _
import json
from datetime import datetime, timedelta
import calendar

def execute(filters=None):
    """
    ERPNext Report Framework entry point
    Returns columns and data for the report
    """
    
    # Get filters with proper defaults
    company = filters.get("company") if filters and filters.get("company") else None
    months_ahead = int(filters.get("months_ahead", 12)) if filters else 12
    include_scenarios = bool(filters.get("include_scenarios", 1)) if filters else True
    forecast_type = filters.get("forecast_type") if filters and filters.get("forecast_type") else None
    account_type = filters.get("account_type") if filters and filters.get("account_type") else None
    confidence_threshold = float(filters.get("confidence_threshold", 0)) if filters else 0
    from_date = filters.get("from_date") if filters else None
    to_date = filters.get("to_date") if filters else None
    
    # Get the main report data with filters
    report_result = generate_cashflow_projection_report(
        company=company, 
        months_ahead=months_ahead, 
        include_scenarios=include_scenarios,
        from_date=from_date,
        to_date=to_date
    )
    
    if not report_result.get("success"):
        frappe.throw(_("Error generating cash flow report: {0}").format(report_result.get("error")))
    
    data = report_result.get("data", {})
    
    # Apply additional filters
    if forecast_type or account_type or confidence_threshold > 0:
        data = apply_filters(data, forecast_type, account_type, confidence_threshold)
    
    # Define columns for the report
    columns = get_report_columns()
    
    # Convert data to tabular format for ERPNext
    report_data = format_data_for_report(data)
    
    return columns, report_data

def apply_filters(data, forecast_type=None, account_type=None, confidence_threshold=0):
    """Apply additional filters to the report data"""
    
    # Filter monthly projections if needed
    if forecast_type or account_type or confidence_threshold > 0:
        filtered_projections = []
        
        for projection in data.get("monthly_projections", []):
            # Apply confidence threshold filter
            revenue_confidence = projection.get("revenue", {}).get("confidence", 100)
            expense_confidence = projection.get("expenses", {}).get("confidence", 100)
            avg_confidence = (revenue_confidence + expense_confidence) / 2
            
            # Skip if confidence is below threshold
            if confidence_threshold > 0 and avg_confidence < confidence_threshold:
                continue
            
            # Apply account type filter - filter based on source account types
            if account_type and account_type != "All":
                # Check if this projection has data for the selected account type
                projection_accounts = projection.get("source_accounts", [])
                if projection_accounts:
                    # Filter accounts by type
                    matching_accounts = [acc for acc in projection_accounts if acc.get("account_type") == account_type]
                    if not matching_accounts:
                        continue  # Skip this projection if no matching account types
                    
                    # Recalculate amounts based on filtered accounts only
                    filtered_revenue = sum(acc.get("credit", 0) for acc in matching_accounts)
                    filtered_expense = sum(acc.get("debit", 0) for acc in matching_accounts)
                    
                    projection["revenue"] = {"amount": filtered_revenue, "confidence": revenue_confidence}
                    projection["expenses"] = {"amount": filtered_expense, "confidence": expense_confidence}
                    projection["net_cashflow"] = {"amount": filtered_revenue - filtered_expense, "confidence": avg_confidence}
            
            # Apply forecast type filter by adjusting amounts
            if forecast_type:
                if forecast_type == "Revenue":
                    # Show only revenue (set expenses to 0)
                    projection["expenses"] = {"amount": 0, "confidence": 0}
                    projection["net_cashflow"] = projection.get("revenue", {"amount": 0})
                elif forecast_type == "Expense":
                    # Show only expenses (set revenue to 0)  
                    projection["revenue"] = {"amount": 0, "confidence": 0}
                    expense_amount = projection.get("expenses", {}).get("amount", 0)
                    projection["net_cashflow"] = {"amount": -abs(expense_amount), "confidence": expense_confidence}
                # For "Cash Flow", show both (no changes needed)
            
            filtered_projections.append(projection)
        
        data["monthly_projections"] = filtered_projections
        
        # Update summary information
        if "summary" in data:
            total_projections = len(filtered_projections)
            data["summary"]["filtered_projections"] = total_projections
            data["summary"]["filter_applied"] = True
            if forecast_type:
                data["summary"]["forecast_type_filter"] = forecast_type
            if confidence_threshold > 0:
                data["summary"]["confidence_threshold"] = confidence_threshold
    
    return data

def get_report_columns():
    """Define columns for the ERPNext report"""
    return [
        {
            "fieldname": "period",
            "label": _("Period"),
            "fieldtype": "Data",
            "width": 120
        },
        {
            "fieldname": "opening_balance",
            "label": _("Opening Balance"),
            "fieldtype": "Currency",
            "width": 150
        },
        {
            "fieldname": "cash_inflow",
            "label": _("Cash Inflow"),
            "fieldtype": "Currency",
            "width": 150
        },
        {
            "fieldname": "cash_outflow",
            "label": _("Cash Outflow"),
            "fieldtype": "Currency",
            "width": 150
        },
        {
            "fieldname": "net_cashflow",
            "label": _("Net Cash Flow"),
            "fieldtype": "Currency",
            "width": 150
        },
        {
            "fieldname": "closing_balance",
            "label": _("Closing Balance"),
            "fieldtype": "Currency",
            "width": 150
        },
        {
            "fieldname": "cumulative_balance",
            "label": _("Cumulative Balance"),
            "fieldtype": "Currency",
            "width": 150
        },
        {
            "fieldname": "risk_level",
            "label": _("Risk Level"),
            "fieldtype": "Data",
            "width": 100
        }
    ]

def format_data_for_report(data):
    """Convert report data to tabular format for ERPNext"""
    
    report_rows = []
    monthly_projections = data.get("monthly_projections", [])
    
    # Add current position row
    current = data.get("current_position", {})
    current_balance = current.get("current_balance", {})
    # Extract total balance if it's an object, otherwise use as is
    current_balance_amount = current_balance.get("total_balance", 0) if isinstance(current_balance, dict) else current_balance
    
    report_rows.append({
        "period": "Current",
        "opening_balance": current_balance_amount,
        "cash_inflow": 0,
        "cash_outflow": 0,
        "net_cashflow": 0,
        "closing_balance": current_balance_amount,
        "cumulative_balance": current_balance_amount,
        "risk_level": "Current"
    })
    
    # Track running balance for proper calculation
    running_balance = current_balance_amount
    
    # Add monthly projection rows
    # monthly_projections is a list, not a dict with projections key
    for projection in monthly_projections:
        revenue_amount = projection.get("revenue", {}).get("amount", 0)
        expense_amount = projection.get("expenses", {}).get("amount", 0)
        net_amount = projection.get("net_cashflow", {}).get("amount", 0)
        
        # Calculate proper opening and closing balances
        opening_balance = running_balance
        closing_balance = opening_balance + net_amount
        running_balance = closing_balance
        
        report_rows.append({
            "period": projection.get("month"),
            "opening_balance": opening_balance,
            "cash_inflow": revenue_amount,
            "cash_outflow": abs(expense_amount),  # Make expenses positive for display
            "net_cashflow": net_amount,
            "closing_balance": closing_balance,
            "cumulative_balance": projection.get("cumulative_cashflow", closing_balance),
            "risk_level": "High" if closing_balance < 0 else ("Medium" if closing_balance < 100000 else "Low")
        })
    
    return report_rows

@frappe.whitelist()
def generate_cashflow_projection_report(company=None, months_ahead=12, include_scenarios=True, from_date=None, to_date=None):
    """Generate comprehensive cash flow projection report"""
    
    try:
        # Get current balance
        current_balance = get_current_balance(company)
        
        # Get monthly projections with date filters
        monthly_projections = get_monthly_projections(company, months_ahead, from_date, to_date)
        
        # Get cash flow breakdown
        cashflow_breakdown = get_cashflow_breakdown(company, months_ahead)
        
        # Get scenario analysis
        scenarios = get_scenario_analysis(company, months_ahead) if include_scenarios else {}
        
        # Get risk assessment
        risk_assessment = get_cashflow_risk_assessment(monthly_projections)
        
        # Get seasonal patterns
        seasonal_analysis = get_seasonal_patterns(company)
        
        # Get key metrics
        key_metrics = calculate_key_metrics(monthly_projections, current_balance)
        
        report_data = {
            "report_title": "Cash Flow Projection Report",
            "generated_at": frappe.utils.now(),
            "company": company or "All Companies",
            "projection_period": f"{months_ahead} months",
            "current_position": {
                "current_balance": current_balance,
                "as_of_date": frappe.utils.nowdate(),
                "account_summary": get_account_summary(company)
            },
            "monthly_projections": monthly_projections,
            "cashflow_breakdown": cashflow_breakdown,
            "scenario_analysis": scenarios,
            "risk_assessment": risk_assessment,
            "seasonal_analysis": seasonal_analysis,
            "key_metrics": key_metrics,
            "recommendations": generate_cashflow_recommendations(monthly_projections, risk_assessment)
        }
        
        return {
            "success": True,
            "data": report_data
        }
        
    except Exception as e:
        frappe.log_error(f"Cash Flow Projection Report Error: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }

def get_current_balance(company=None):
    """Get current cash balance from all cash accounts"""
    
    cash_accounts = frappe.get_all("Account",
                                  filters={
                                      "account_type": ["in", ["Bank", "Cash"]],
                                      "is_group": 0,
                                      "company": company
                                  } if company else {
                                      "account_type": ["in", ["Bank", "Cash"]],
                                      "is_group": 0
                                  },
                                  fields=["name", "account_name"])
    
    total_balance = 0
    account_details = []
    
    for account in cash_accounts:
        # Get balance from GL Entry
        balance = frappe.db.sql("""
            SELECT 
                SUM(debit) - SUM(credit) as balance
            FROM `tabGL Entry`
            WHERE account = %s 
            AND is_cancelled = 0
            AND posting_date <= %s
        """, (account.name, frappe.utils.nowdate()))[0][0] or 0
        
        total_balance += balance
        account_details.append({
            "account": account.name,
            "account_name": account.account_name,
            "balance": balance
        })
    
    return {
        "total_balance": total_balance,
        "account_breakdown": account_details,
        "currency": frappe.defaults.get_global_default("currency") or "INR"
    }

def get_monthly_projections(company=None, months_ahead=12, from_date=None, to_date=None):
    """Get monthly cash flow projections"""
    
    projections = []
    
    # Use from_date if provided, otherwise start from today
    start_date = frappe.utils.getdate(from_date) if from_date else frappe.utils.getdate()
    
    # If to_date is provided, calculate months_ahead based on date range
    if to_date:
        end_date = frappe.utils.getdate(to_date)
        months_ahead = max(1, frappe.utils.date_diff(end_date, start_date) // 30)
    
    for month_offset in range(months_ahead):
        projection_date = frappe.utils.add_months(start_date, month_offset)
        month_start = frappe.utils.get_first_day(projection_date)
        month_end = frappe.utils.get_last_day(projection_date)
        
        # Skip if beyond to_date
        if to_date and month_start > frappe.utils.getdate(to_date):
            break
        
        # Get revenue forecasts for this month
        revenue_forecast = frappe.db.sql("""
            SELECT 
                SUM(predicted_amount) as total_revenue,
                AVG(confidence_score) as avg_confidence
            FROM `tabAI Financial Forecast`
            WHERE forecast_type = 'Revenue'
            AND forecast_start_date <= %(month_end)s
            AND forecast_end_date >= %(month_start)s
            {}
        """.format("AND company = %(company)s" if company else ""),
        {"company": company, "month_start": month_start, "month_end": month_end}, as_dict=True)[0]
        
        # Get expense forecasts for this month
        expense_forecast = frappe.db.sql("""
            SELECT 
                SUM(predicted_amount) as total_expenses,
                AVG(confidence_score) as avg_confidence
            FROM `tabAI Financial Forecast`
            WHERE forecast_type = 'Expense'
            AND forecast_start_date <= %(month_end)s
            AND forecast_end_date >= %(month_start)s
            {}
        """.format("AND company = %(company)s" if company else ""),
        {"company": company, "month_start": month_start, "month_end": month_end}, as_dict=True)[0]
        
        # Get cash flow forecasts for this month
        cashflow_forecast = frappe.db.sql("""
            SELECT 
                SUM(predicted_amount) as net_cashflow,
                AVG(confidence_score) as avg_confidence,
                COUNT(*) as forecast_count
            FROM `tabAI Financial Forecast`
            WHERE forecast_type = 'Cash Flow'
            AND forecast_start_date <= %(month_end)s
            AND forecast_end_date >= %(month_start)s
            {}
        """.format("AND company = %(company)s" if company else ""),
        {"company": company, "month_start": month_start, "month_end": month_end}, as_dict=True)[0]
        
        total_revenue = revenue_forecast.total_revenue or 0
        total_expenses = expense_forecast.total_expenses or 0
        net_cashflow = cashflow_forecast.net_cashflow or (total_revenue - total_expenses)
        
        projections.append({
            "month": projection_date.strftime("%Y-%m"),
            "month_name": calendar.month_name[projection_date.month],
            "year": projection_date.year,
            "revenue": {
                "amount": total_revenue,
                "confidence": revenue_forecast.avg_confidence or 0
            },
            "expenses": {
                "amount": total_expenses,
                "confidence": expense_forecast.avg_confidence or 0
            },
            "net_cashflow": {
                "amount": net_cashflow,
                "confidence": cashflow_forecast.avg_confidence or 0
            },
            "forecast_count": cashflow_forecast.forecast_count or 0,
            "cumulative_cashflow": 0  # Will be calculated later
        })
    
    # Calculate cumulative cash flow
    cumulative = 0
    for projection in projections:
        cumulative += projection["net_cashflow"]["amount"]
        projection["cumulative_cashflow"] = cumulative
    
    return projections

def get_cashflow_breakdown(company=None, months_ahead=12):
    """Get detailed cash flow breakdown by category"""
    
    # Get inflow categories
    inflow_categories = frappe.db.sql("""
        SELECT 
            account_type,
            SUM(predicted_amount) as total_amount,
            AVG(confidence_score) as avg_confidence,
            COUNT(*) as forecast_count
        FROM `tabAI Financial Forecast`
        WHERE forecast_type IN ('Revenue', 'Cash Flow')
        AND predicted_amount > 0
        AND forecast_start_date >= %(start_date)s
        AND forecast_start_date <= %(end_date)s
        {}
        GROUP BY account_type
        ORDER BY total_amount DESC
    """.format("AND company = %(company)s" if company else ""),
    {
        "company": company,
        "start_date": frappe.utils.nowdate(),
        "end_date": frappe.utils.add_months(frappe.utils.nowdate(), months_ahead)
    }, as_dict=True)
    
    # Get outflow categories
    outflow_categories = frappe.db.sql("""
        SELECT 
            account_type,
            SUM(ABS(predicted_amount)) as total_amount,
            AVG(confidence_score) as avg_confidence,
            COUNT(*) as forecast_count
        FROM `tabAI Financial Forecast`
        WHERE forecast_type IN ('Expense', 'Cash Flow')
        AND (predicted_amount < 0 OR forecast_type = 'Expense')
        AND forecast_start_date >= %(start_date)s
        AND forecast_start_date <= %(end_date)s
        {}
        GROUP BY account_type
        ORDER BY total_amount DESC
    """.format("AND company = %(company)s" if company else ""),
    {
        "company": company,
        "start_date": frappe.utils.nowdate(),
        "end_date": frappe.utils.add_months(frappe.utils.nowdate(), months_ahead)
    }, as_dict=True)
    
    # Calculate totals
    total_inflows = sum(cat.total_amount for cat in inflow_categories)
    total_outflows = sum(cat.total_amount for cat in outflow_categories)
    
    return {
        "inflows": {
            "categories": inflow_categories,
            "total": total_inflows,
            "avg_confidence": sum(cat.avg_confidence for cat in inflow_categories) / len(inflow_categories) if inflow_categories else 0
        },
        "outflows": {
            "categories": outflow_categories,
            "total": total_outflows,
            "avg_confidence": sum(cat.avg_confidence for cat in outflow_categories) / len(outflow_categories) if outflow_categories else 0
        },
        "net_flow": total_inflows - total_outflows
    }

def get_scenario_analysis(company=None, months_ahead=12):
    """Generate scenario analysis (optimistic, pessimistic, most likely)"""
    
    scenarios = {}
    
    # Base scenario (most likely) - use existing forecasts
    base_projections = get_monthly_projections(company, months_ahead)
    
    scenarios["realistic"] = {
        "net_cashflow": sum(p["net_cashflow"]["amount"] for p in base_projections),
        "probability": 60,
        "assumptions": "Based on current AI forecasts and historical patterns"
    }
    
    # Optimistic scenario (+20% revenue, -10% expenses)
    optimistic_projections = []
    for projection in base_projections:
        opt_projection = projection.copy()
        opt_projection["revenue"]["amount"] *= 1.2
        opt_projection["expenses"]["amount"] *= 0.9
        opt_projection["net_cashflow"]["amount"] = opt_projection["revenue"]["amount"] - opt_projection["expenses"]["amount"]
        optimistic_projections.append(opt_projection)
    
    scenarios["optimistic"] = {
        "net_cashflow": sum(p["net_cashflow"]["amount"] for p in optimistic_projections),
        "probability": 25,
        "assumptions": "20% higher revenue, 10% lower expenses, favorable market conditions"
    }
    
    # Pessimistic scenario (-15% revenue, +15% expenses)
    pessimistic_projections = []
    for projection in base_projections:
        pess_projection = projection.copy()
        pess_projection["revenue"]["amount"] *= 0.85
        pess_projection["expenses"]["amount"] *= 1.15
        pess_projection["net_cashflow"]["amount"] = pess_projection["revenue"]["amount"] - pess_projection["expenses"]["amount"]
        pessimistic_projections.append(pess_projection)
    
    scenarios["pessimistic"] = {
        "net_cashflow": sum(p["net_cashflow"]["amount"] for p in pessimistic_projections),
        "probability": 15,
        "assumptions": "15% lower revenue, 15% higher expenses, market downturn"
    }
    
    return scenarios

def get_cashflow_risk_assessment(monthly_projections):
    """Assess cash flow risks"""
    
    risks = []
    risk_score = 0
    
    # Check for negative cash flow months
    negative_months = [p for p in monthly_projections if p["net_cashflow"]["amount"] < 0]
    if negative_months:
        risks.append({
            "type": "Negative Cash Flow",
            "severity": "High" if len(negative_months) > 3 else "Medium",
            "description": f"{len(negative_months)} months with negative cash flow projected",
            "affected_months": [p["month"] for p in negative_months],
            "mitigation": "Consider credit facilities or expense reduction"
        })
        risk_score += len(negative_months) * 10
    
    # Check for low confidence forecasts
    low_confidence = [p for p in monthly_projections if p["net_cashflow"]["confidence"] < 70]
    if low_confidence:
        risks.append({
            "type": "Forecast Uncertainty",
            "severity": "Medium",
            "description": f"{len(low_confidence)} months with low forecast confidence (<70%)",
            "affected_months": [p["month"] for p in low_confidence],
            "mitigation": "Improve data quality and forecasting models"
        })
        risk_score += len(low_confidence) * 5
    
    # Check for high volatility
    if len(monthly_projections) > 1:
        amounts = [p["net_cashflow"]["amount"] for p in monthly_projections]
        avg_amount = sum(amounts) / len(amounts)
        volatility = sum(abs(amount - avg_amount) for amount in amounts) / len(amounts)
        volatility_pct = (volatility / abs(avg_amount)) * 100 if avg_amount != 0 else 0
        
        if volatility_pct > 50:
            risks.append({
                "type": "High Volatility",
                "severity": "Medium",
                "description": f"Cash flow volatility of {volatility_pct:.1f}% detected",
                "volatility_percentage": volatility_pct,
                "mitigation": "Establish cash reserves for stability"
            })
            risk_score += 15
    
    # Determine overall risk level
    if risk_score < 20:
        risk_level = "Low"
    elif risk_score < 50:
        risk_level = "Medium"
    else:
        risk_level = "High"
    
    return {
        "overall_risk_level": risk_level,
        "risk_score": min(100, risk_score),
        "identified_risks": risks,
        "risk_summary": f"{len(risks)} risks identified with {risk_level.lower()} overall risk level"
    }

def get_seasonal_patterns(company=None):
    """Analyze seasonal patterns in cash flow"""
    
    # Get historical data for seasonal analysis
    historical_data = frappe.db.sql("""
        SELECT 
            MONTH(forecast_start_date) as month,
            AVG(predicted_amount) as avg_amount,
            COUNT(*) as forecast_count
        FROM `tabAI Financial Forecast`
        WHERE forecast_type = 'Cash Flow'
        AND creation >= DATE_SUB(NOW(), INTERVAL 24 MONTH)
        {}
        GROUP BY MONTH(forecast_start_date)
        ORDER BY month
    """.format("AND company = %(company)s" if company else ""),
    {"company": company}, as_dict=True)
    
    if not historical_data:
        return {"available": False, "message": "Insufficient historical data for seasonal analysis"}
    
    # Calculate seasonal factors
    overall_avg = sum(h.avg_amount for h in historical_data) / len(historical_data)
    seasonal_factors = []
    
    for month_data in historical_data:
        factor = month_data.avg_amount / overall_avg if overall_avg != 0 else 1
        seasonal_factors.append({
            "month": month_data.month,
            "month_name": calendar.month_name[month_data.month],
            "seasonal_factor": factor,
            "avg_amount": month_data.avg_amount,
            "forecast_count": month_data.forecast_count
        })
    
    # Identify peak and low seasons
    sorted_factors = sorted(seasonal_factors, key=lambda x: x["seasonal_factor"], reverse=True)
    peak_months = sorted_factors[:3]
    low_months = sorted_factors[-3:]
    
    return {
        "available": True,
        "seasonal_factors": seasonal_factors,
        "peak_months": [m["month_name"] for m in peak_months],
        "low_months": [m["month_name"] for m in low_months],
        "seasonality_strength": max(seasonal_factors, key=lambda x: x["seasonal_factor"])["seasonal_factor"] - 
                               min(seasonal_factors, key=lambda x: x["seasonal_factor"])["seasonal_factor"]
    }

def calculate_key_metrics(monthly_projections, current_balance):
    """Calculate key cash flow metrics"""
    
    if not monthly_projections:
        return {}
    
    # Basic calculations
    total_projected_inflow = sum(p["revenue"]["amount"] for p in monthly_projections)
    total_projected_outflow = sum(p["expenses"]["amount"] for p in monthly_projections)
    net_cashflow = total_projected_inflow - total_projected_outflow
    
    # Average monthly cash flow
    avg_monthly_cashflow = net_cashflow / len(monthly_projections)
    
    # Cash runway (months until cash runs out)
    current_cash = current_balance.get("total_balance", 0)
    if avg_monthly_cashflow < 0:
        cash_runway_months = current_cash / abs(avg_monthly_cashflow)
    else:
        cash_runway_months = float('inf')  # Positive cash flow
    
    # Minimum balance projection
    running_balance = current_cash
    min_balance = current_cash
    min_balance_month = None
    
    for projection in monthly_projections:
        running_balance += projection["net_cashflow"]["amount"]
        if running_balance < min_balance:
            min_balance = running_balance
            min_balance_month = projection["month"]
    
    # Cash conversion efficiency
    if total_projected_inflow > 0:
        conversion_efficiency = (net_cashflow / total_projected_inflow) * 100
    else:
        conversion_efficiency = 0
    
    # Forecast reliability score
    avg_confidence = sum(p["net_cashflow"]["confidence"] for p in monthly_projections) / len(monthly_projections)
    
    return {
        "total_projected_inflow": total_projected_inflow,
        "total_projected_outflow": total_projected_outflow,
        "net_cashflow_projection": net_cashflow,
        "avg_monthly_cashflow": avg_monthly_cashflow,
        "cash_runway_months": cash_runway_months if cash_runway_months != float('inf') else None,
        "minimum_balance": {
            "amount": min_balance,
            "month": min_balance_month
        },
        "cash_conversion_efficiency": conversion_efficiency,
        "forecast_reliability_score": avg_confidence,
        "liquidity_status": "Healthy" if min_balance > 0 else "At Risk"
    }

def get_account_summary(company=None):
    """Get summary of all accounts contributing to cash flow"""
    
    accounts = frappe.db.sql("""
        SELECT 
            aff.account,
            aff.account_name,
            aff.account_type,
            COUNT(*) as forecast_count,
            SUM(aff.predicted_amount) as total_predicted,
            AVG(aff.confidence_score) as avg_confidence,
            MAX(aff.forecast_end_date) as latest_forecast
        FROM `tabAI Financial Forecast` aff
        WHERE aff.forecast_start_date >= %(start_date)s
        {}
        GROUP BY aff.account, aff.account_name, aff.account_type
        ORDER BY ABS(SUM(aff.predicted_amount)) DESC
        LIMIT 20
    """.format("AND aff.company = %(company)s" if company else ""),
    {
        "company": company,
        "start_date": frappe.utils.nowdate()
    }, as_dict=True)
    
    return accounts

def generate_cashflow_recommendations(monthly_projections, risk_assessment):
    """Generate actionable cash flow recommendations"""
    
    recommendations = []
    
    # Check for negative cash flow
    negative_months = [p for p in monthly_projections if p["net_cashflow"]["amount"] < 0]
    if negative_months:
        recommendations.append({
            "priority": "High",
            "title": "Address Negative Cash Flow Periods",
            "description": f"Prepare for {len(negative_months)} months with negative cash flow",
            "action_items": [
                "Establish credit line or overdraft facility",
                "Consider accelerating receivables collection",
                "Review and postpone non-essential expenses",
                "Negotiate extended payment terms with suppliers"
            ],
            "timeline": "Immediate",
            "impact": "Critical for business continuity"
        })
    
    # Low confidence forecasts
    low_confidence = [p for p in monthly_projections if p["net_cashflow"]["confidence"] < 70]
    if len(low_confidence) > 3:
        recommendations.append({
            "priority": "Medium",
            "title": "Improve Forecast Accuracy",
            "description": f"{len(low_confidence)} months have low forecast confidence",
            "action_items": [
                "Review and clean historical data",
                "Implement more frequent data updates",
                "Consider alternative forecasting models",
                "Add more data sources for validation"
            ],
            "timeline": "1-2 months",
            "impact": "Better planning and risk management"
        })
    
    # High risk assessment
    if risk_assessment["overall_risk_level"] in ["High", "Critical"]:
        recommendations.append({
            "priority": "Critical",
            "title": "Implement Risk Mitigation Strategy",
            "description": f"Overall cash flow risk level is {risk_assessment['overall_risk_level']}",
            "action_items": [
                "Create emergency cash reserve fund",
                "Diversify revenue streams",
                "Implement stricter expense controls",
                "Regular cash flow monitoring (weekly)"
            ],
            "timeline": "Immediate",
            "impact": "Reduce financial risk exposure"
        })
    
    # Positive recommendations for good cash flow
    positive_months = [p for p in monthly_projections if p["net_cashflow"]["amount"] > 0]
    if len(positive_months) > len(monthly_projections) * 0.8:  # 80% positive months
        recommendations.append({
            "priority": "Low",
            "title": "Optimize Excess Cash",
            "description": "Strong positive cash flow projected",
            "action_items": [
                "Consider short-term investments",
                "Accelerate debt repayments",
                "Invest in business growth opportunities",
                "Build strategic cash reserves"
            ],
            "timeline": "3-6 months",
            "impact": "Maximize returns on excess cash"
        })
    
    return recommendations

@frappe.whitelist()
def export_cashflow_report(company=None, months_ahead=12, format="excel"):
    """Export cash flow report in specified format"""
    
    report_data = generate_cashflow_projection_report(company, months_ahead, True)
    
    if not report_data["success"]:
        return report_data
    
    if format == "excel":
        return export_to_excel(report_data["data"])
    elif format == "pdf":
        return export_to_pdf(report_data["data"])
    else:
        return report_data

def export_to_excel(data):
    """Export report data to Excel format"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add headers and data for summary
        summary_data = [
            ["Cash Flow Projection Summary", ""],
            ["Report Generated", data["generated_at"]],
            ["Company", data["company"]],
            ["Projection Period", data["projection_period"]],
            ["", ""],
            ["Current Position", ""],
            ["Current Balance", data["current_position"]["current_balance"]["total_balance"]],
            ["Currency", data["current_position"]["current_balance"]["currency"]],
            ["", ""],
            ["Key Metrics", ""],
            ["Net Cash Flow Projection", data["key_metrics"]["net_cashflow_projection"]],
            ["Average Monthly Cash Flow", data["key_metrics"]["avg_monthly_cashflow"]],
            ["Minimum Balance", data["key_metrics"]["minimum_balance"]["amount"]],
            ["Liquidity Status", data["key_metrics"]["liquidity_status"]]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                if col_num == 1 and value and value != "":
                    cell.font = Font(bold=True)
        
        # Monthly Projections sheet
        ws_monthly = wb.create_sheet("Monthly Projections")
        
        monthly_headers = ["Month", "Revenue", "Expenses", "Net Cash Flow", "Cumulative", "Confidence"]
        for col_num, header in enumerate(monthly_headers, 1):
            cell = ws_monthly.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, projection in enumerate(data["monthly_projections"], 2):
            ws_monthly.cell(row=row_num, column=1, value=projection["month"])
            ws_monthly.cell(row=row_num, column=2, value=projection["revenue"]["amount"])
            ws_monthly.cell(row=row_num, column=3, value=projection["expenses"]["amount"])
            ws_monthly.cell(row=row_num, column=4, value=projection["net_cashflow"]["amount"])
            ws_monthly.cell(row=row_num, column=5, value=projection["cumulative_cashflow"])
            ws_monthly.cell(row=row_num, column=6, value=f"{projection['net_cashflow']['confidence']:.1f}%")
        
        # Risk Assessment sheet
        ws_risks = wb.create_sheet("Risk Assessment")
        
        risk_data = [
            ["Risk Assessment", ""],
            ["Overall Risk Level", data["risk_assessment"]["overall_risk_level"]],
            ["Risk Score", f"{data['risk_assessment']['risk_score']}/100"],
            ["", ""],
            ["Identified Risks", ""]
        ]
        
        for risk in data["risk_assessment"]["identified_risks"]:
            risk_data.extend([
                [f"Risk: {risk['type']}", ""],
                ["Severity", risk["severity"]],
                ["Description", risk["description"]],
                ["Mitigation", risk["mitigation"]],
                ["", ""]
            ])
        
        for row_num, row_data in enumerate(risk_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_risks.cell(row=row_num, column=col_num, value=value)
                if col_num == 1 and ":" not in str(value) and value != "":
                    cell.font = Font(bold=True)
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            "success": True,
            "content": output.getvalue(),
            "filename": f"cashflow_projection_{frappe.utils.today()}.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
    except ImportError:
        # Fallback to simple CSV if openpyxl not available
        return export_to_csv(data)
    except Exception as e:
        frappe.log_error(f"Excel export error: {str(e)}")
        return {"success": False, "error": str(e)}

def export_to_csv(data):
    """Export report data to CSV format"""
    
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Write monthly projections
    writer.writerow(["Month", "Revenue", "Expenses", "Net Cash Flow", "Cumulative", "Confidence"])
    
    for projection in data["monthly_projections"]:
        writer.writerow([
            projection["month"],
            projection["revenue"]["amount"],
            projection["expenses"]["amount"],
            projection["net_cashflow"]["amount"],
            projection["cumulative_cashflow"],
            f"{projection['net_cashflow']['confidence']:.1f}%"
        ])
    
    return {
        "success": True,
        "content": output.getvalue(),
        "filename": f"cashflow_projection_{frappe.utils.today()}.csv",
        "content_type": "text/csv"
    }

def export_to_pdf(data):
    """Export report data to PDF format"""
    
    html_content = generate_cashflow_html_report(data)
    
    try:
        pdf_content = frappe.utils.get_pdf(html_content)
        return {
            "success": True,
            "content": pdf_content,
            "filename": f"cashflow_projection_{frappe.utils.today()}.pdf",
            "content_type": "application/pdf"
        }
    except Exception as e:
        frappe.log_error(f"PDF export error: {str(e)}")
        return {"success": False, "error": str(e)}

def generate_cashflow_html_report(data):
    """Generate HTML version of cash flow report"""
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Cash Flow Projection Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
            .header {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
            .section {{ margin-bottom: 30px; }}
            .metric-card {{ display: inline-block; margin: 10px; padding: 15px; border: 1px solid #ddd; border-radius: 5px; text-align: center; }}
            table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            .positive {{ color: #28a745; }}
            .negative {{ color: #dc3545; }}
            .high-risk {{ background-color: #f8d7da; }}
            .medium-risk {{ background-color: #fff3cd; }}
            .low-risk {{ background-color: #d4edda; }}
            .chart-placeholder {{ height: 300px; background: #f8f9fa; display: flex; align-items: center; justify-content: center; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Cash Flow Projection Report</h1>
            <p><strong>Company:</strong> {data['company']}</p>
            <p><strong>Generated:</strong> {data['generated_at']}</p>
            <p><strong>Projection Period:</strong> {data['projection_period']}</p>
        </div>
        
        <div class="section">
            <h2>Executive Summary</h2>
            <div class="metric-card">
                <h3>₹{data['current_position']['current_balance']['total_balance']:,.2f}</h3>
                <p>Current Balance</p>
            </div>
            <div class="metric-card">
                <h3 class="{'positive' if data['key_metrics']['net_cashflow_projection'] > 0 else 'negative'}">
                    ₹{data['key_metrics']['net_cashflow_projection']:,.2f}
                </h3>
                <p>Net Cash Flow (Projected)</p>
            </div>
            <div class="metric-card">
                <h3>₹{data['key_metrics']['avg_monthly_cashflow']:,.2f}</h3>
                <p>Average Monthly Cash Flow</p>
            </div>
            <div class="metric-card">
                <h3 class="{data['risk_assessment']['overall_risk_level'].lower()}-risk">
                    {data['risk_assessment']['overall_risk_level']}
                </h3>
                <p>Risk Level</p>
            </div>
        </div>
        
        <div class="section">
            <h2>Monthly Projections</h2>
            <table>
                <thead>
                    <tr>
                        <th>Month</th>
                        <th>Revenue</th>
                        <th>Expenses</th>
                        <th>Net Cash Flow</th>
                        <th>Cumulative</th>
                        <th>Confidence</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for projection in data["monthly_projections"]:
        net_class = "positive" if projection["net_cashflow"]["amount"] > 0 else "negative"
        html += f"""
                    <tr>
                        <td>{projection['month']}</td>
                        <td>₹{projection['revenue']['amount']:,.2f}</td>
                        <td>₹{projection['expenses']['amount']:,.2f}</td>
                        <td class="{net_class}">₹{projection['net_cashflow']['amount']:,.2f}</td>
                        <td>₹{projection['cumulative_cashflow']:,.2f}</td>
                        <td>{projection['net_cashflow']['confidence']:.1f}%</td>
                    </tr>
        """
    
    html += """
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2>Risk Assessment</h2>
            <p><strong>Overall Risk Level:</strong> <span class="{}-risk">{}</span></p>
            <p><strong>Risk Score:</strong> {}/100</p>
            
            <h3>Identified Risks</h3>
    """.format(
        data['risk_assessment']['overall_risk_level'].lower(),
        data['risk_assessment']['overall_risk_level'],
        data['risk_assessment']['risk_score']
    )
    
    for risk in data['risk_assessment']['identified_risks']:
        html += f"""
            <div class="{risk['severity'].lower()}-risk" style="padding: 10px; margin: 10px 0; border-radius: 5px;">
                <h4>{risk['type']} - {risk['severity']} Risk</h4>
                <p>{risk['description']}</p>
                <p><strong>Mitigation:</strong> {risk['mitigation']}</p>
            </div>
        """
    
    html += """
        </div>
        
        <div class="section">
            <h2>Recommendations</h2>
    """
    
    for rec in data['recommendations']:
        priority_class = rec['priority'].lower()
        html += f"""
            <div class="{priority_class}-risk" style="padding: 15px; margin: 10px 0; border-radius: 5px;">
                <h4>{rec['title']} ({rec['priority']} Priority)</h4>
                <p>{rec['description']}</p>
                <ul>
        """
        for action in rec['action_items']:
            html += f"<li>{action}</li>"
        html += f"""
                </ul>
                <p><strong>Timeline:</strong> {rec['timeline']} | <strong>Impact:</strong> {rec['impact']}</p>
            </div>
        """
    
    html += """
        </div>
    </body>
    </html>
    """
    
    return html

@frappe.whitelist()
def export_report(filters=None, format='excel'):
    """Export cash flow report in specified format"""
    try:
        # Get report data
        columns, data = execute(filters)
        
        if format == 'excel':
            return export_to_excel(columns, data)
        elif format == 'pdf':
            return export_to_pdf(columns, data)
        else:
            return {"success": False, "error": "Unsupported format"}
            
    except Exception as e:
        frappe.log_error(f"Export Report Error: {str(e)}")
        return {"success": False, "error": str(e)}

@frappe.whitelist()
def email_report(filters=None, recipients=None, subject=None, message=None):
    """Email cash flow report to specified recipients"""
    try:
        # Get report data
        columns, data = execute(filters)
        
        # Generate Excel attachment
        excel_data = export_to_excel(columns, data)
        
        if not excel_data.get("success"):
            return {"success": False, "error": "Failed to generate Excel report"}
        
        # Send email with attachment
        frappe.sendmail(
            recipients=recipients.split(",") if isinstance(recipients, str) else recipients,
            subject=subject or "Cash Flow Projection Report",
            message=message or "Please find attached the cash flow projection report.",
            attachments=[{
                "fname": excel_data.get("filename", "cashflow_report.xlsx"),
                "fcontent": excel_data.get("content")
            }]
        )
        
        return {"success": True, "message": "Report emailed successfully"}
        
    except Exception as e:
        frappe.log_error(f"Email Report Error: {str(e)}")
        return {"success": False, "error": str(e)}

@frappe.whitelist()
def get_scenario_analysis_data(filters=None):
    """Get detailed scenario analysis data"""
    try:
        # Handle filters parameter - could be string, dict, or None
        if isinstance(filters, str):
            import json
            try:
                filters = json.loads(filters)
            except:
                filters = {}
        elif filters is None:
            filters = {}
        
        company = filters.get("company") if filters else None
        months_ahead = filters.get("months_ahead", 12) if filters else 12
        
        scenarios = get_scenario_analysis(company, months_ahead)
        
        return {"success": True, "data": scenarios}
        
    except Exception as e:
        frappe.log_error(f"Scenario Analysis Error: {str(e)}")
        return {"success": False, "error": str(e)}

@frappe.whitelist()
def create_dashboard_charts(filters=None):
    """Create dashboard charts for cash flow data"""
    try:
        # Handle filters parameter - could be string, dict, or None
        if isinstance(filters, str):
            import json
            try:
                filters = json.loads(filters)
            except:
                filters = {}
        elif filters is None:
            filters = {}
            
        # Get report data
        columns, data = execute(filters)
        
        # Calculate summary metrics
        total_inflow = sum(row.get("cash_inflow", 0) for row in data if row.get("period") != "Current")
        total_outflow = sum(row.get("cash_outflow", 0) for row in data if row.get("period") != "Current")
        net_flow = total_inflow - total_outflow
        
        # Create chart data
        chart_data = {
            "total_inflow": total_inflow,
            "total_outflow": total_outflow,
            "net_flow": net_flow,
            "period": f"Next {len([row for row in data if row.get('period') != 'Current'])} months",
            "data_points": len(data),
            "confidence": "Medium",
            "labels": [row.get("period") for row in data if row.get("period") != "Current"],
            "datasets": [
                {
                    "name": "Cash Inflow",
                    "values": [row.get("cash_inflow", 0) for row in data if row.get("period") != "Current"]
                },
                {
                    "name": "Cash Outflow", 
                    "values": [row.get("cash_outflow", 0) for row in data if row.get("period") != "Current"]
                },
                {
                    "name": "Net Cash Flow",
                    "values": [row.get("net_cashflow", 0) for row in data if row.get("period") != "Current"]
                }
            ]
        }
        
        return {"success": True, "chart_data": chart_data, "message": "Dashboard charts data generated"}
        
    except Exception as e:
        frappe.log_error(f"Dashboard Charts Error: {str(e)}")
        return {"success": False, "error": str(e)}

def export_to_excel(columns, data):
    """Export data to Excel format"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Projection"
        
        # Add headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column.get("label"))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, column in enumerate(columns, 1):
                fieldname = column.get("fieldname")
                value = row_data.get(fieldname, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        from io import BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return {
            "success": True,
            "content": excel_file.getvalue(),
            "filename": f"cashflow_projection_report_{frappe.utils.nowdate()}.xlsx"
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}

def export_to_pdf(columns, data):
    """Export data to PDF format"""
    try:
        # Simple HTML to PDF conversion
        html_content = f"""
        <html>
        <head>
            <title>Cash Flow Projection Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #366092; color: white; }}
                .currency {{ text-align: right; }}
            </style>
        </head>
        <body>
            <h1>Cash Flow Projection Report</h1>
            <table>
                <thead>
                    <tr>
                        {"".join([f"<th>{col.get('label')}</th>" for col in columns])}
                    </tr>
                </thead>
                <tbody>
        """
        
        for row in data:
            html_content += "<tr>"
            for col in columns:
                fieldname = col.get("fieldname")
                value = row.get(fieldname, "")
                if col.get("fieldtype") == "Currency":
                    value = frappe.format(value, {"fieldtype": "Currency"})
                    html_content += f'<td class="currency">{value}</td>'
                else:
                    html_content += f"<td>{value}</td>"
            html_content += "</tr>"
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        return {
            "success": True,
            "content": html_content,
            "filename": f"cashflow_projection_report_{frappe.utils.nowdate()}.pdf"
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}