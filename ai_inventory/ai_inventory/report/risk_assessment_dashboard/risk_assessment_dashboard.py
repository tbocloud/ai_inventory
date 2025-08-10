# risk_assessment_dashboard.py
import frappe
from frappe import _
import json
from datetime import datetime, timedelta
import statistics

def execute(filters=None):
    """Main execute function for ERPNext report"""
    columns = get_columns()
    data = get_data(filters)
    return columns, data

def get_columns():
    """Define report columns"""
    return [
        {
            "fieldname": "risk_category",
            "label": _("Risk Category"),
            "fieldtype": "Data",
            "width": 150
        },
        {
            "fieldname": "risk_type",
            "label": _("Risk Type"),
            "fieldtype": "Data",
            "width": 200
        },
        {
            "fieldname": "risk_description",
            "label": _("Description"),
            "fieldtype": "Data",
            "width": 500
        },
        {
            "fieldname": "probability",
            "label": _("Probability %"),
            "fieldtype": "Percent",
            "width": 120
        },
        {
            "fieldname": "impact_score",
            "label": _("Impact Score"),
            "fieldtype": "Float",
            "width": 120
        },
        {
            "fieldname": "risk_score",
            "label": _("Risk Score"),
            "fieldtype": "Float",
            "width": 120
        },
        {
            "fieldname": "risk_level",
            "label": _("Risk Level"),
            "fieldtype": "Data",
            "width": 100
        },
        {
            "fieldname": "mitigation_status",
            "label": _("Mitigation Status"),
            "fieldtype": "Data",
            "width": 150
        },
        {
            "fieldname": "trend",
            "label": _("Trend"),
            "fieldtype": "Data",
            "width": 150
        }
    ]

def get_data(filters):
    """Get report data"""
    try:
        # Extract filter values
        company = filters.get("company") if filters else None
        period_months = int(filters.get("period_months", 6)) if filters and filters.get("period_months") else 6
        risk_category = filters.get("risk_category", "All") if filters else "All"
        risk_level = filters.get("risk_level", "All") if filters else "All"
        
        # Calculate date range based on period_months
        from frappe.utils import add_months, today, getdate
        end_date = getdate(today())
        start_date = add_months(end_date, -period_months)
        
        # Generate risk assessment data based on available financial forecast data
        report_rows = []
        
        # Get financial forecast data to assess risks
        forecast_data = frappe.db.sql("""
            SELECT 
                COUNT(*) as forecast_count,
                AVG(confidence_score) as avg_confidence,
                AVG(predicted_amount) as avg_predicted,
                STDDEV(predicted_amount) as volatility,
                prediction_model,
                account_type,
                forecast_period_days
            FROM `tabAI Financial Forecast`
            WHERE (company = %(company)s OR %(company)s IS NULL)
            GROUP BY prediction_model, account_type, forecast_period_days
            ORDER BY volatility DESC
            LIMIT 20
        """, {"company": company}, as_dict=True)
        
        # Financial Risks
        if risk_category in ("All", "Financial"):
            # Filter forecast data based on selected period (convert months to days) if data exists
            if forecast_data:
                target_period_days = period_months * 30
                relevant_forecasts = [f for f in forecast_data if abs(f.forecast_period_days - target_period_days) <= 30]
                
                # If no forecasts match the period, use all available data with period adjustment
                if not relevant_forecasts:
                    relevant_forecasts = forecast_data
            else:
                relevant_forecasts = []
            
            # Period adjustment factor - longer periods have higher risk exposure
            period_multiplier = 1 + (period_months - 6) * 0.1  # Base is 6 months
            
            # Cash Flow Risk
            if relevant_forecasts:
                cash_flow_volatility = next((f.volatility for f in relevant_forecasts if 'Cash' in str(f.account_type)), 0) or 0
                cash_flow_risk_score = min((cash_flow_volatility / 100000) * period_multiplier, 10)  # Normalize to 1-10 scale
            else:
                # Fallback values when no forecast data is available
                cash_flow_risk_score = 3.5 * period_multiplier
            
            report_rows.append({
                "risk_category": "Financial",
                "risk_type": "Cash Flow Volatility",
                "risk_description": f"High volatility in cash flow predictions over {period_months} months indicating potential liquidity issues",
                "probability": min(cash_flow_risk_score * 10, 100),
                "impact_score": 8.5,
                "risk_score": cash_flow_risk_score,
                "risk_level": get_risk_level_text(cash_flow_risk_score),
                "mitigation_status": "Active" if cash_flow_risk_score > 5 else "Monitoring",
                "trend": "Increasing" if cash_flow_risk_score > 6 else "Stable"
            })
            
            # Revenue Concentration Risk
            if relevant_forecasts:
                avg_confidence = sum(f.avg_confidence for f in relevant_forecasts) / len(relevant_forecasts)
                concentration_risk = ((100 - avg_confidence) / 10) * period_multiplier  # Convert to 1-10 scale with period adjustment
            else:
                # Fallback values when no forecast data is available
                avg_confidence = 75
                concentration_risk = 2.5 * period_multiplier
            
            report_rows.append({
                "risk_category": "Financial",
                "risk_type": "Revenue Concentration",
                "risk_description": f"Dependency on limited revenue sources over {period_months}-month forecast period",
                "probability": 100 - avg_confidence,
                "impact_score": 7.0,
                "risk_score": concentration_risk,
                "risk_level": get_risk_level_text(concentration_risk),
                "mitigation_status": "Under Review" if concentration_risk > 4 else "Controlled",
                "trend": "Stable"
            })
            
            # Credit Risk
            credit_risk_score = 4.2
            report_rows.append({
                "risk_category": "Financial",
                "risk_type": "Credit Risk",
                "risk_description": "Potential losses from customer payment defaults",
                "probability": 15.0,
                "impact_score": 6.5,
                "risk_score": credit_risk_score,
                "risk_level": get_risk_level_text(credit_risk_score),
                "mitigation_status": "Controlled",
                "trend": "Decreasing"
            })
        
        # Operational Risks
        if risk_category in ("All", "Operational"):
            report_rows.extend([
                {
                    "risk_category": "Operational",
                    "risk_type": "Supply Chain Disruption",
                    "risk_description": "Potential disruptions in supply chain affecting operations",
                    "probability": 25.0,
                    "impact_score": 8.0,
                    "risk_score": 6.5,
                    "risk_level": "High",
                    "mitigation_status": "Active",
                    "trend": "Increasing"
                },
                {
                    "risk_category": "Operational",
                    "risk_type": "Technology System Failure",
                    "risk_description": "Risk of critical system failures affecting operations",
                    "probability": 12.0,
                    "impact_score": 7.5,
                    "risk_score": 4.8,
                    "risk_level": "Medium",
                    "mitigation_status": "Controlled",
                    "trend": "Stable"
                },
                {
                    "risk_category": "Operational",
                    "risk_type": "Key Personnel Risk",
                    "risk_description": "Risk of losing critical personnel and knowledge",
                    "probability": 18.0,
                    "impact_score": 6.0,
                    "risk_score": 5.2,
                    "risk_level": "Medium",
                    "mitigation_status": "Under Review",
                    "trend": "Stable"
                }
            ])
        
        # Market Risks
        if risk_category in ("All", "Market"):
            # Calculate market volatility from forecast data
            market_volatility = statistics.stdev([f.avg_predicted for f in forecast_data]) if len(forecast_data) > 1 else 0
            market_risk_score = min(market_volatility / 200000, 10)
            
            report_rows.extend([
                {
                    "risk_category": "Market",
                    "risk_type": "Market Volatility",
                    "risk_description": "Risk from market fluctuations affecting business performance",
                    "probability": min(market_risk_score * 8, 100),
                    "impact_score": 7.8,
                    "risk_score": market_risk_score,
                    "risk_level": get_risk_level_text(market_risk_score),
                    "mitigation_status": "Monitoring",
                    "trend": "Increasing" if market_risk_score > 5 else "Stable"
                },
                {
                    "risk_category": "Market",
                    "risk_type": "Competitive Risk",
                    "risk_description": "Risk from increased competition and market share loss",
                    "probability": 35.0,
                    "impact_score": 6.5,
                    "risk_score": 5.8,
                    "risk_level": "Medium",
                    "mitigation_status": "Active",
                    "trend": "Stable"
                },
                {
                    "risk_category": "Market",
                    "risk_type": "Regulatory Risk",
                    "risk_description": "Risk from changing regulations and compliance requirements",
                    "probability": 20.0,
                    "impact_score": 7.0,
                    "risk_score": 4.5,
                    "risk_level": "Medium",
                    "mitigation_status": "Controlled",
                    "trend": "Stable"
                }
            ])
        
        # Compliance Risks
        if risk_category in ("All", "Compliance"):
            report_rows.extend([
                {
                    "risk_category": "Compliance",
                    "risk_type": "Data Privacy Risk",
                    "risk_description": "Risk of data breaches and privacy violations",
                    "probability": 8.0,
                    "impact_score": 9.0,
                    "risk_score": 3.8,
                    "risk_level": "Low",
                    "mitigation_status": "Controlled",
                    "trend": "Decreasing"
                },
                {
                    "risk_category": "Compliance",
                    "risk_type": "Financial Reporting Risk",
                    "risk_description": "Risk of errors in financial reporting and compliance",
                    "probability": 5.0,
                    "impact_score": 8.5,
                    "risk_score": 2.9,
                    "risk_level": "Low",
                    "mitigation_status": "Controlled",
                    "trend": "Stable"
                }
            ])
        
        # Apply risk_level filter if specified
        if risk_level != "All":
            report_rows = [row for row in report_rows if row.get("risk_level") == risk_level]
        
        return report_rows
        
    except Exception as e:
        frappe.log_error(f"Risk Assessment Dashboard Error: {str(e)}")
        return []

def get_risk_level_text(risk_score):
    """Convert numeric risk score to text level"""
    if risk_score >= 7:
        return "Critical"
    elif risk_score >= 5:
        return "High"
    elif risk_score >= 3:
        return "Medium"
    else:
        return "Low"

@frappe.whitelist()
def generate_risk_assessment_dashboard(company=None, period_months=6):
    """Generate comprehensive risk assessment dashboard"""
    
    try:
        # Get financial risks
        financial_risks = assess_financial_risks(company, period_months)
        
        # Get operational risks  
        operational_risks = assess_operational_risks(company, period_months)
        
        # Get market risks
        market_risks = assess_market_risks(company, period_months)
        
        # Get risk metrics
        risk_metrics = calculate_risk_metrics(company, period_months)
        
        # Get mitigation strategies
        mitigation_strategies = get_mitigation_strategies(financial_risks, operational_risks, market_risks)
        
        # Calculate overall risk score
        overall_risk_score = calculate_overall_risk_score(financial_risks, operational_risks, market_risks)
        
        # Generate risk insights
        risk_insights = generate_risk_insights(financial_risks, operational_risks, market_risks, risk_metrics)
        
        report_data = {
            "report_title": "Risk Assessment Dashboard",
            "generated_at": frappe.utils.now(),
            "company": company or "All Companies",
            "assessment_period": f"{period_months} months",
            "overall_risk_score": overall_risk_score,
            "risk_level": get_risk_level(overall_risk_score),
            "financial_risks": financial_risks,
            "operational_risks": operational_risks, 
            "market_risks": market_risks,
            "risk_metrics": risk_metrics,
            "mitigation_strategies": mitigation_strategies,
            "risk_insights": risk_insights,
            "risk_trend": get_risk_trend(company, period_months)
        }
        
        return {
            "success": True,
            "data": report_data
        }
        
    except Exception as e:
        frappe.log_error(f"Risk Assessment Dashboard Error: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }

def assess_financial_risks(company=None, period_months=6):
    """Assess financial risks"""
    
    risks = []
    
    # Cash flow risk
    cash_flow_risk = assess_cash_flow_risk(company, period_months)
    if cash_flow_risk:
        risks.append(cash_flow_risk)
    
    # Liquidity risk
    liquidity_risk = assess_liquidity_risk(company)
    if liquidity_risk:
        risks.append(liquidity_risk)
    
    # Credit risk
    credit_risk = assess_credit_risk(company)
    if credit_risk:
        risks.append(credit_risk)
    
    # Forecast accuracy risk
    accuracy_risk = assess_forecast_accuracy_risk(company)
    if accuracy_risk:
        risks.append(accuracy_risk)
    
    return risks

def assess_cash_flow_risk(company=None, period_months=6):
    """Assess cash flow risks"""
    
    # Get cash flow forecasts
    negative_months = frappe.db.sql("""
        SELECT 
            COUNT(*) as negative_count,
            MIN(predicted_amount) as worst_month,
            AVG(predicted_amount) as avg_flow
        FROM `tabAI Financial Forecast`
        WHERE forecast_type = 'Cash Flow'
        AND predicted_amount < 0
        AND forecast_start_date >= CURDATE()
        AND forecast_start_date <= DATE_ADD(CURDATE(), INTERVAL %(period_months)s MONTH)
        {}
    """.format("AND company = %(company)s" if company else ""),
    {"company": company, "period_months": period_months}, as_dict=True)[0]
    
    if negative_months.negative_count > 0:
        severity = "Critical" if negative_months.negative_count > 3 else "High" if negative_months.negative_count > 1 else "Medium"
        
        return {
            "risk_type": "Cash Flow Risk",
            "category": "Financial",
            "severity": severity,
            "probability": min(90, negative_months.negative_count * 20),
            "impact": "High",
            "description": f"Forecasted negative cash flow for {negative_months.negative_count} months",
            "worst_case_scenario": f"₹{abs(negative_months.worst_month):,.2f} negative in worst month",
            "current_exposure": abs(negative_months.avg_flow) if negative_months.avg_flow < 0 else 0,
            "trend": "Increasing",
            "mitigation_urgency": "Immediate" if severity == "Critical" else "High",
            "related_accounts": get_related_cash_accounts(company),
            "risk_factors": [
                "Seasonal demand variations",
                "Payment collection delays", 
                "High fixed costs",
                "Inventory investment requirements"
            ]
        }
    
    return None

def assess_liquidity_risk(company=None):
    """Assess liquidity risks"""
    
    # Get current liquid assets
    liquid_assets = frappe.db.sql("""
        SELECT 
            SUM(CASE WHEN account_type IN ('Bank', 'Cash') THEN predicted_amount ELSE 0 END) as cash_equivalents,
            SUM(CASE WHEN account_type = 'Receivable' THEN predicted_amount ELSE 0 END) as receivables
        FROM `tabAI Financial Forecast`
        WHERE forecast_type = 'Balance Sheet'
        AND forecast_start_date <= CURDATE()
        {}
        ORDER BY forecast_start_date DESC
        LIMIT 1
    """.format("AND company = %(company)s" if company else ""),
    {"company": company}, as_dict=True)[0]
    
    # Get current liabilities
    current_liabilities = frappe.db.sql("""
        SELECT 
            SUM(predicted_amount) as total_liabilities
        FROM `tabAI Financial Forecast`
        WHERE account_type = 'Payable'
        AND forecast_start_date <= CURDATE()
        {}
        ORDER BY forecast_start_date DESC
        LIMIT 1
    """.format("AND company = %(company)s" if company else ""),
    {"company": company}, as_dict=True)[0]
    
    cash_equivalents = liquid_assets.cash_equivalents or 0
    receivables = liquid_assets.receivables or 0
    total_liabilities = current_liabilities.total_liabilities or 1
    
    # Calculate liquidity ratios
    current_ratio = (cash_equivalents + receivables) / total_liabilities
    quick_ratio = cash_equivalents / total_liabilities
    
    if current_ratio < 1.5 or quick_ratio < 1.0:
        severity = "High" if current_ratio < 1.0 else "Medium"
        
        return {
            "risk_type": "Liquidity Risk",
            "category": "Financial", 
            "severity": severity,
            "probability": 70,
            "impact": "High",
            "description": f"Low liquidity ratios detected (Current: {current_ratio:.2f}, Quick: {quick_ratio:.2f})",
            "current_exposure": total_liabilities - (cash_equivalents + receivables),
            "trend": "Stable",
            "mitigation_urgency": "High" if severity == "High" else "Medium",
            "metrics": {
                "current_ratio": round(current_ratio, 2),
                "quick_ratio": round(quick_ratio, 2),
                "cash_equivalents": cash_equivalents,
                "receivables": receivables,
                "current_liabilities": total_liabilities
            },
            "risk_factors": [
                "Insufficient cash reserves",
                "High accounts payable",
                "Slow receivables collection",
                "Limited credit facilities"
            ]
        }
    
    return None

def assess_credit_risk(company=None):
    """Assess credit and counterparty risks"""
    
    # Get receivables aging analysis
    overdue_receivables = frappe.db.sql("""
        SELECT 
            SUM(predicted_amount) as total_receivables,
            AVG(confidence_score) as avg_confidence
        FROM `tabAI Financial Forecast`
        WHERE account_type = 'Receivable'
        AND forecast_start_date <= CURDATE()
        {}
    """.format("AND company = %(company)s" if company else ""),
    {"company": company}, as_dict=True)[0]
    
    total_receivables = overdue_receivables.total_receivables or 0
    confidence = overdue_receivables.avg_confidence or 100
    
    # Estimate credit risk based on receivables and confidence
    estimated_bad_debt = total_receivables * (100 - confidence) / 100
    
    if estimated_bad_debt > total_receivables * 0.05:  # >5% bad debt risk
        severity = "High" if estimated_bad_debt > total_receivables * 0.1 else "Medium"
        
        return {
            "risk_type": "Credit Risk",
            "category": "Financial",
            "severity": severity,
            "probability": 100 - confidence,
            "impact": "Medium",
            "description": f"Potential bad debt exposure of ₹{estimated_bad_debt:,.2f}",
            "current_exposure": estimated_bad_debt,
            "trend": "Stable",
            "mitigation_urgency": "Medium",
            "metrics": {
                "total_receivables": total_receivables,
                "estimated_bad_debt": estimated_bad_debt,
                "bad_debt_percentage": (estimated_bad_debt / max(total_receivables, 1)) * 100,
                "collection_confidence": confidence
            },
            "risk_factors": [
                "Customer payment delays",
                "Economic downturn impact",
                "Concentration of major customers",
                "Inadequate credit screening"
            ]
        }
    
    return None

def assess_forecast_accuracy_risk(company=None):
    """Assess risks from forecast inaccuracy"""
    
    # Get recent forecast accuracy
    accuracy_data = frappe.db.sql("""
        SELECT 
            AVG(accuracy_score) as avg_accuracy,
            STDDEV(accuracy_score) as accuracy_volatility,
            COUNT(*) as total_forecasts
        FROM `tabAI Forecast Accuracy` aa
        JOIN `tabAI Financial Forecast` aff ON aa.original_forecast_id = aff.name
        WHERE aa.evaluation_date >= DATE_SUB(NOW(), INTERVAL 90 DAY)
        {}
    """.format("AND aff.company = %(company)s" if company else ""),
    {"company": company}, as_dict=True)[0]
    
    avg_accuracy = accuracy_data.avg_accuracy or 80
    accuracy_volatility = accuracy_data.accuracy_volatility or 10
    
    if avg_accuracy < 75 or accuracy_volatility > 20:
        severity = "High" if avg_accuracy < 60 else "Medium"
        
        return {
            "risk_type": "Forecast Accuracy Risk", 
            "category": "Financial",
            "severity": severity,
            "probability": 100 - avg_accuracy,
            "impact": "High",
            "description": f"Low forecast accuracy ({avg_accuracy:.1f}%) affecting planning reliability",
            "current_exposure": 0,  # Indirect exposure
            "trend": "Stable",
            "mitigation_urgency": "High",
            "metrics": {
                "average_accuracy": round(avg_accuracy, 1),
                "accuracy_volatility": round(accuracy_volatility, 1),
                "total_forecasts_evaluated": accuracy_data.total_forecasts
            },
            "risk_factors": [
                "Poor data quality",
                "Model limitations",
                "Market volatility",
                "External factor changes"
            ]
        }
    
    return None

def assess_operational_risks(company=None, period_months=6):
    """Assess operational risks"""
    
    risks = []
    
    # Capacity utilization risk
    capacity_risk = assess_capacity_risk(company)
    if capacity_risk:
        risks.append(capacity_risk)
    
    # Supply chain risk
    supply_chain_risk = assess_supply_chain_risk(company)
    if supply_chain_risk:
        risks.append(supply_chain_risk)
    
    # Quality risk
    quality_risk = assess_quality_risk(company)
    if quality_risk:
        risks.append(quality_risk)
    
    # Technology risk
    technology_risk = assess_technology_risk(company)
    if technology_risk:
        risks.append(technology_risk)
    
    return risks

def assess_capacity_risk(company=None):
    """Assess capacity and production risks"""
    
    # Get capacity utilization from manufacturing data
    capacity_data = frappe.db.sql("""
        SELECT 
            AVG(predicted_amount) as avg_production
        FROM `tabAI Financial Forecast`
        WHERE forecast_type = 'Revenue'
        AND (account_name LIKE '%Production%' OR account_name LIKE '%Manufacturing%')
        AND forecast_start_date >= DATE_SUB(CURDATE(), INTERVAL 6 MONTH)
        {}
    """.format("AND company = %(company)s" if company else ""),
    {"company": company}, as_dict=True)[0]
    
    avg_production = capacity_data.avg_production or 1000000
    
    # Simulate capacity utilization (would come from actual production systems)
    estimated_capacity = avg_production * 1.2  # Assume 20% headroom
    utilization = (avg_production / estimated_capacity) * 100
    
    if utilization > 90:
        return {
            "risk_type": "Capacity Constraint Risk",
            "category": "Operational",
            "severity": "High",
            "probability": 80,
            "impact": "High", 
            "description": f"High capacity utilization ({utilization:.1f}%) may limit growth",
            "current_exposure": avg_production * 0.1,  # 10% of production at risk
            "trend": "Increasing",
            "mitigation_urgency": "High",
            "metrics": {
                "current_utilization": round(utilization, 1),
                "avg_production": avg_production,
                "estimated_capacity": estimated_capacity
            },
            "risk_factors": [
                "Equipment bottlenecks",
                "Limited production capacity",
                "Maintenance downtime",
                "Skilled labor shortage"
            ]
        }
    elif utilization < 60:
        return {
            "risk_type": "Underutilization Risk",
            "category": "Operational",
            "severity": "Medium",
            "probability": 70,
            "impact": "Medium",
            "description": f"Low capacity utilization ({utilization:.1f}%) indicates inefficiency",
            "current_exposure": (estimated_capacity - avg_production) * 0.5,  # Cost of idle capacity
            "trend": "Stable", 
            "mitigation_urgency": "Medium",
            "metrics": {
                "current_utilization": round(utilization, 1),
                "unused_capacity": estimated_capacity - avg_production
            },
            "risk_factors": [
                "Reduced demand",
                "Inefficient operations",
                "High fixed costs",
                "Market competition"
            ]
        }
    
    return None

def assess_supply_chain_risk(company=None):
    """Assess supply chain and supplier risks"""
    
    # Get supplier concentration data (placeholder)
    # In reality, this would come from purchase orders and supplier data
    
    return {
        "risk_type": "Supply Chain Risk",
        "category": "Operational",
        "severity": "Medium",
        "probability": 60,
        "impact": "High",
        "description": "Potential supply chain disruptions from key suppliers",
        "current_exposure": 500000,  # Estimated exposure
        "trend": "Stable",
        "mitigation_urgency": "Medium",
        "metrics": {
            "key_supplier_dependency": 65,  # Percentage
            "supplier_diversity_score": 7,   # Out of 10
            "supply_chain_resilience": 75    # Percentage
        },
        "risk_factors": [
            "Single source suppliers",
            "Geographic concentration",
            "Quality control issues",
            "Price volatility",
            "Delivery delays"
        ]
    }

def assess_quality_risk(company=None):
    """Assess quality and compliance risks"""
    
    # Placeholder for quality metrics
    # Would come from quality management systems
    
    return {
        "risk_type": "Quality Risk",
        "category": "Operational",
        "severity": "Low",
        "probability": 30,
        "impact": "Medium",
        "description": "Risk of quality issues affecting customer satisfaction",
        "current_exposure": 200000,
        "trend": "Improving",
        "mitigation_urgency": "Low",
        "metrics": {
            "quality_score": 96.5,
            "defect_rate": 3.5,
            "customer_complaints": 2
        },
        "risk_factors": [
            "Process variations",
            "Raw material quality",
            "Human error",
            "Equipment malfunction"
        ]
    }

def assess_technology_risk(company=None):
    """Assess technology and system risks"""
    
    return {
        "risk_type": "Technology Risk",
        "category": "Operational",
        "severity": "Medium",
        "probability": 40,
        "impact": "Medium",
        "description": "Risk of system failures or technology obsolescence",
        "current_exposure": 300000,
        "trend": "Stable",
        "mitigation_urgency": "Medium",
        "metrics": {
            "system_uptime": 99.2,
            "technology_age": 3.5,  # Years
            "backup_effectiveness": 95
        },
        "risk_factors": [
            "System downtime",
            "Data security breaches", 
            "Technology obsolescence",
            "Integration challenges"
        ]
    }

def assess_market_risks(company=None, period_months=6):
    """Assess market and external risks"""
    
    risks = []
    
    # Demand volatility risk
    demand_risk = assess_demand_risk(company, period_months)
    if demand_risk:
        risks.append(demand_risk)
    
    # Competition risk
    competition_risk = assess_competition_risk(company)
    if competition_risk:
        risks.append(competition_risk)
    
    # Economic risk
    economic_risk = assess_economic_risk(company)
    if economic_risk:
        risks.append(economic_risk)
    
    # Regulatory risk
    regulatory_risk = assess_regulatory_risk(company)
    if regulatory_risk:
        risks.append(regulatory_risk)
    
    return risks

def assess_demand_risk(company=None, period_months=6):
    """Assess demand volatility risks"""
    
    # Get revenue volatility
    revenue_data = frappe.db.sql("""
        SELECT 
            STDDEV(predicted_amount) as revenue_volatility,
            AVG(predicted_amount) as avg_revenue,
            COUNT(*) as data_points
        FROM `tabAI Financial Forecast`
        WHERE forecast_type = 'Revenue'
        AND forecast_start_date >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
        {}
    """.format("AND company = %(company)s" if company else ""),
    {"company": company}, as_dict=True)[0]
    
    revenue_volatility = revenue_data.revenue_volatility or 0
    avg_revenue = revenue_data.avg_revenue or 1000000
    
    volatility_pct = (revenue_volatility / max(avg_revenue, 1)) * 100
    
    if volatility_pct > 20:
        severity = "High" if volatility_pct > 40 else "Medium"
        
        return {
            "risk_type": "Demand Volatility Risk",
            "category": "Market",
            "severity": severity,
            "probability": min(90, volatility_pct * 2),
            "impact": "High",
            "description": f"High demand volatility ({volatility_pct:.1f}%) affecting revenue predictability",
            "current_exposure": revenue_volatility * 2,  # Two standard deviations
            "trend": "Stable",
            "mitigation_urgency": "High" if severity == "High" else "Medium",
            "metrics": {
                "demand_volatility": round(volatility_pct, 1),
                "revenue_std_dev": revenue_volatility,
                "avg_revenue": avg_revenue
            },
            "risk_factors": [
                "Seasonal demand patterns",
                "Economic cycles",
                "Customer concentration",
                "Product lifecycle changes"
            ]
        }
    
    return None

def assess_competition_risk(company=None):
    """Assess competitive risks"""
    
    return {
        "risk_type": "Competition Risk",
        "category": "Market",
        "severity": "Medium",
        "probability": 50,
        "impact": "Medium",
        "description": "Risk of market share loss to competitors",
        "current_exposure": 1000000,  # Estimated revenue at risk
        "trend": "Increasing",
        "mitigation_urgency": "Medium",
        "metrics": {
            "market_share": 25,  # Percentage
            "competitive_position": 3,  # Rank
            "price_competitiveness": 85  # Score
        },
        "risk_factors": [
            "New market entrants",
            "Price competition",
            "Technology disruption",
            "Customer switching"
        ]
    }

def assess_economic_risk(company=None):
    """Assess macroeconomic risks"""
    
    return {
        "risk_type": "Economic Risk",
        "category": "Market",
        "severity": "Medium",
        "probability": 40,
        "impact": "High",
        "description": "Risk from economic downturn or inflation",
        "current_exposure": 2000000,
        "trend": "Stable",
        "mitigation_urgency": "Low",
        "metrics": {
            "economic_sensitivity": 70,  # How sensitive business is to economy
            "inflation_impact": 15,      # Percentage impact from inflation
            "recession_resilience": 60   # Score out of 100
        },
        "risk_factors": [
            "Economic recession",
            "Inflation pressure",
            "Interest rate changes",
            "Currency fluctuation"
        ]
    }

def assess_regulatory_risk(company=None):
    """Assess regulatory and compliance risks"""
    
    return {
        "risk_type": "Regulatory Risk",
        "category": "Market",
        "severity": "Low",
        "probability": 25,
        "impact": "Medium",
        "description": "Risk from regulatory changes or compliance issues",
        "current_exposure": 500000,
        "trend": "Stable",
        "mitigation_urgency": "Low",
        "metrics": {
            "compliance_score": 92,
            "regulatory_changes": 2,  # Number per year
            "audit_results": 95       # Score
        },
        "risk_factors": [
            "New regulations",
            "Compliance failures",
            "License renewals",
            "Environmental standards"
        ]
    }

def calculate_risk_metrics(company=None, period_months=6):
    """Calculate key risk metrics"""
    
    # Value at Risk calculation (simplified)
    total_exposure = frappe.db.sql("""
        SELECT 
            SUM(ABS(predicted_amount)) as total_exposure
        FROM `tabAI Financial Forecast`
        WHERE forecast_start_date >= CURDATE()
        AND forecast_start_date <= DATE_ADD(CURDATE(), INTERVAL %(period_months)s MONTH)
        {}
    """.format("AND company = %(company)s" if company else ""),
    {"company": company, "period_months": period_months}, as_dict=True)[0]
    
    exposure = total_exposure.total_exposure or 10000000
    
    # Calculate VaR (simplified - 95% confidence, 1% of exposure)
    var_95 = exposure * 0.01
    var_99 = exposure * 0.005
    
    return {
        "value_at_risk_95": var_95,
        "value_at_risk_99": var_99,
        "total_exposure": exposure,
        "risk_adjusted_return": 15.2,  # Placeholder
        "diversification_ratio": 0.75,
        "stress_test_result": 85,  # Score out of 100
        "risk_appetite": "Moderate",
        "risk_tolerance": 1000000,  # Maximum acceptable loss
        "early_warning_indicators": [
            "Cash flow negative trend",
            "Accuracy below 70%",
            "Capacity utilization >90%"
        ]
    }

def calculate_overall_risk_score(financial_risks, operational_risks, market_risks):
    """Calculate overall risk score"""
    
    all_risks = financial_risks + operational_risks + market_risks
    
    if not all_risks:
        return 20  # Low risk if no risks identified
    
    # Weight risks by severity and probability
    severity_weights = {"Critical": 100, "High": 75, "Medium": 50, "Low": 25}
    
    total_weighted_score = 0
    total_weight = 0
    
    for risk in all_risks:
        severity_weight = severity_weights.get(risk["severity"], 50)
        probability = risk.get("probability", 50)
        
        risk_score = (severity_weight * probability) / 100
        total_weighted_score += risk_score
        total_weight += 1
    
    # Normalize to 0-100 scale
    if total_weight > 0:
        avg_risk_score = total_weighted_score / total_weight
        return min(100, max(0, avg_risk_score))
    
    return 20

def get_risk_level(risk_score):
    """Convert risk score to risk level"""
    
    if risk_score >= 80:
        return "Critical"
    elif risk_score >= 60:
        return "High" 
    elif risk_score >= 40:
        return "Medium"
    else:
        return "Low"

def get_mitigation_strategies(financial_risks, operational_risks, market_risks):
    """Generate mitigation strategies for identified risks"""
    
    strategies = []
    all_risks = financial_risks + operational_risks + market_risks
    
    # Group strategies by risk category
    for risk in all_risks:
        if risk["severity"] in ["Critical", "High"]:
            strategy = generate_mitigation_strategy(risk)
            if strategy:
                strategies.append(strategy)
    
    return strategies

def generate_mitigation_strategy(risk):
    """Generate specific mitigation strategy for a risk"""
    
    risk_type = risk["risk_type"]
    
    strategies_map = {
        "Cash Flow Risk": {
            "strategy": "Cash Flow Management",
            "actions": [
                "Establish credit line facility",
                "Accelerate receivables collection",
                "Negotiate extended payment terms with suppliers",
                "Implement cash flow forecasting system"
            ],
            "timeline": "1-3 months",
            "cost": "Medium",
            "effectiveness": "High"
        },
        "Liquidity Risk": {
            "strategy": "Liquidity Enhancement", 
            "actions": [
                "Maintain minimum cash reserves",
                "Diversify funding sources",
                "Optimize working capital",
                "Establish standby credit facilities"
            ],
            "timeline": "1-2 months",
            "cost": "Low",
            "effectiveness": "High"
        },
        "Capacity Constraint Risk": {
            "strategy": "Capacity Expansion",
            "actions": [
                "Invest in additional equipment",
                "Optimize production scheduling",
                "Consider outsourcing peak demand",
                "Implement lean manufacturing"
            ],
            "timeline": "6-12 months", 
            "cost": "High",
            "effectiveness": "High"
        },
        "Supply Chain Risk": {
            "strategy": "Supply Chain Diversification",
            "actions": [
                "Develop alternative suppliers",
                "Increase safety stock levels",
                "Implement supplier monitoring",
                "Geographic diversification"
            ],
            "timeline": "3-6 months",
            "cost": "Medium",
            "effectiveness": "Medium"
        }
    }
    
    base_strategy = strategies_map.get(risk_type)
    if base_strategy:
        return {
            "risk_type": risk_type,
            "risk_severity": risk["severity"],
            "strategy_name": base_strategy["strategy"],
            "recommended_actions": base_strategy["actions"],
            "implementation_timeline": base_strategy["timeline"],
            "estimated_cost": base_strategy["cost"],
            "expected_effectiveness": base_strategy["effectiveness"],
            "priority": "High" if risk["severity"] in ["Critical", "High"] else "Medium",
            "responsible_department": get_responsible_department(risk["category"]),
            "success_metrics": get_success_metrics(risk_type)
        }
    
    return None

def get_responsible_department(risk_category):
    """Get responsible department for risk category"""
    
    department_map = {
        "Financial": "Finance Department",
        "Operational": "Operations Department", 
        "Market": "Sales & Marketing Department"
    }
    
    return department_map.get(risk_category, "Management")

def get_success_metrics(risk_type):
    """Get success metrics for risk mitigation"""
    
    metrics_map = {
        "Cash Flow Risk": ["Positive monthly cash flow", "Cash reserves >30 days expenses"],
        "Liquidity Risk": ["Current ratio >1.5", "Quick ratio >1.0"],
        "Capacity Constraint Risk": ["Utilization 70-85%", "On-time delivery >95%"],
        "Supply Chain Risk": ["Supplier diversity >3", "Stockout incidents <2%"]
    }
    
    return metrics_map.get(risk_type, ["Risk exposure reduced by 50%"])

def generate_risk_insights(financial_risks, operational_risks, market_risks, risk_metrics):
    """Generate insights from risk analysis"""
    
    insights = []
    all_risks = financial_risks + operational_risks + market_risks
    
    # Risk concentration insight
    risk_categories = {}
    for risk in all_risks:
        category = risk["category"]
        if category not in risk_categories:
            risk_categories[category] = 0
        risk_categories[category] += 1
    
    if risk_categories:
        max_category = max(risk_categories, key=risk_categories.get)
        max_count = risk_categories[max_category]
        
        if max_count > 2:
            insights.append({
                "type": "warning",
                "title": f"Risk Concentration in {max_category}",
                "description": f"{max_count} risks identified in {max_category.lower()} category",
                "recommendation": f"Focus mitigation efforts on {max_category.lower()} risks"
            })
    
    # High severity risks insight
    critical_risks = [r for r in all_risks if r["severity"] == "Critical"]
    high_risks = [r for r in all_risks if r["severity"] == "High"]
    
    if critical_risks:
        insights.append({
            "type": "critical",
            "title": f"{len(critical_risks)} Critical Risk(s) Identified",
            "description": "Immediate action required for critical risks",
            "recommendation": "Implement emergency mitigation measures"
        })
    
    if len(high_risks) > 3:
        insights.append({
            "type": "warning",
            "title": "Multiple High-Severity Risks",
            "description": f"{len(high_risks)} high-severity risks require attention",
            "recommendation": "Prioritize high-impact mitigation strategies"
        })
    
    # Risk trend insight
    increasing_risks = [r for r in all_risks if r.get("trend") == "Increasing"]
    if len(increasing_risks) > 2:
        insights.append({
            "type": "warning",
            "title": "Increasing Risk Trends",
            "description": f"{len(increasing_risks)} risks showing increasing trend",
            "recommendation": "Monitor risk indicators more frequently"
        })
    
    return insights

def get_risk_trend(company=None, period_months=6):
    """Calculate risk trend over time"""
    
    # This would analyze historical risk assessments
    # For now, return a placeholder trend
    
    return {
        "direction": "Stable",
        "change_percentage": 2.1,
        "trend_analysis": "Risk levels have remained relatively stable with slight increase in operational risks"
    }

def get_related_cash_accounts(company=None):
    """Get cash-related accounts for risk analysis"""
    
    accounts = frappe.db.sql("""
        SELECT DISTINCT account, account_name
        FROM `tabAI Financial Forecast`
        WHERE account_type IN ('Bank', 'Cash', 'Receivable', 'Payable')
        {}
        LIMIT 10
    """.format("AND company = %(company)s" if company else ""),
    {"company": company}, as_dict=True)
    
    return [acc.account_name for acc in accounts]

@frappe.whitelist()
def export_risk_assessment(company=None, period_months=6, format="excel"):
    """Export risk assessment in specified format"""
    
    report_data = generate_risk_assessment_dashboard(company, period_months)
    
    if not report_data["success"]:
        return report_data
    
    if format == "excel":
        return export_risk_to_excel(report_data["data"])
    elif format == "pdf":
        return export_risk_to_pdf(report_data["data"])
    else:
        return report_data

def export_risk_to_excel(data):
    """Export risk assessment to Excel"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Risk Summary"
        
        summary_data = [
            ["Risk Assessment Dashboard", ""],
            ["Generated", data["generated_at"]],
            ["Company", data["company"]],
            ["Assessment Period", data["assessment_period"]],
            ["", ""],
            ["Overall Risk Score", data["overall_risk_score"]],
            ["Risk Level", data["risk_level"]],
            ["Total Risks Identified", len(data["financial_risks"]) + len(data["operational_risks"]) + len(data["market_risks"])],
            ["Value at Risk (95%)", data["risk_metrics"]["value_at_risk_95"]]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                if col_num == 1 and value:
                    cell.font = Font(bold=True)
        
        # Risk Details sheet
        ws_risks = wb.create_sheet("Risk Details")
        
        risk_headers = ["Risk Type", "Category", "Severity", "Probability", "Impact", "Description", "Mitigation Urgency"]
        for col_num, header in enumerate(risk_headers, 1):
            cell = ws_risks.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        all_risks = data["financial_risks"] + data["operational_risks"] + data["market_risks"]
        for row_num, risk in enumerate(all_risks, 2):
            ws_risks.cell(row=row_num, column=1, value=risk["risk_type"])
            ws_risks.cell(row=row_num, column=2, value=risk["category"])
            ws_risks.cell(row=row_num, column=3, value=risk["severity"])
            ws_risks.cell(row=row_num, column=4, value=f"{risk['probability']}%")
            ws_risks.cell(row=row_num, column=5, value=risk["impact"])
            ws_risks.cell(row=row_num, column=6, value=risk["description"])
            ws_risks.cell(row=row_num, column=7, value=risk["mitigation_urgency"])
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            "success": True,
            "content": output.getvalue(),
            "filename": f"risk_assessment_{frappe.utils.today()}.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
    except Exception as e:
        frappe.log_error(f"Risk Excel export error: {str(e)}")
        return {"success": False, "error": str(e)}

def export_risk_to_pdf(data):
    """Export risk assessment to PDF"""
    
    html_content = generate_risk_html_report(data)
    
    try:
        pdf_content = frappe.utils.get_pdf(html_content)
        return {
            "success": True,
            "content": pdf_content,
            "filename": f"risk_assessment_{frappe.utils.today()}.pdf",
            "content_type": "application/pdf"
        }
    except Exception as e:
        frappe.log_error(f"Risk PDF export error: {str(e)}")
        return {"success": False, "error": str(e)}

def generate_risk_html_report(data):
    """Generate HTML version of risk assessment"""
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Risk Assessment Dashboard</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            .header {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
            .risk-score {{ text-align: center; margin: 20px 0; }}
            .risk-critical {{ background: #f8d7da; }}
            .risk-high {{ background: #fff3cd; }}
            .risk-medium {{ background: #d1ecf1; }}
            .risk-low {{ background: #d4edda; }}
            table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
            th {{ background-color: #f2f2f2; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Risk Assessment Dashboard</h1>
            <p><strong>Company:</strong> {data['company']}</p>
            <p><strong>Generated:</strong> {data['generated_at']}</p>
            <p><strong>Assessment Period:</strong> {data['assessment_period']}</p>
        </div>
        
        <div class="risk-score">
            <h2>Overall Risk Score: {data['overall_risk_score']}/100</h2>
            <h3>Risk Level: {data['risk_level']}</h3>
        </div>
        
        <h2>Risk Summary</h2>
        <table>
            <tr>
                <th>Risk Category</th>
                <th>Number of Risks</th>
                <th>Highest Severity</th>
            </tr>
            <tr>
                <td>Financial</td>
                <td>{len(data['financial_risks'])}</td>
                <td>{max([r['severity'] for r in data['financial_risks']], default='None')}</td>
            </tr>
            <tr>
                <td>Operational</td>
                <td>{len(data['operational_risks'])}</td>
                <td>{max([r['severity'] for r in data['operational_risks']], default='None')}</td>
            </tr>
            <tr>
                <td>Market</td>
                <td>{len(data['market_risks'])}</td>
                <td>{max([r['severity'] for r in data['market_risks']], default='None')}</td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    return html