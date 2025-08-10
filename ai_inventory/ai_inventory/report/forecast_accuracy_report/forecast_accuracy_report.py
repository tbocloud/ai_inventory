# forecast_accuracy_report.py
import frappe
from frappe import _
import json
from datetime import datetime, timedelta
import statistics

def execute(filters=None):
    """Main execute function for ERPNext report"""
    columns = get_columns()
    data = get_data(filters)
    return columns, data

def get_columns():
    """Define report columns"""
    return [
        {
            "fieldname": "forecast_id",
            "label": _("Forecast ID"),
            "fieldtype": "Link",
            "options": "AI Financial Forecast",
            "width": 200
        },
        {
            "fieldname": "account",
            "label": _("Account"),
            "fieldtype": "Data",
            "width": 400
        },
        {
            "fieldname": "predicted_amount",
            "label": _("Predicted Amount"),
            "fieldtype": "Currency",
            "width": 150
        },
        {
            "fieldname": "actual_amount",
            "label": _("Actual Amount"),
            "fieldtype": "Currency",
            "width": 150
        },
        {
            "fieldname": "accuracy_percentage",
            "label": _("Accuracy %"),
            "fieldtype": "Percent",
            "width": 150
        },
        {
            "fieldname": "confidence_score",
            "label": _("Confidence"),
            "fieldtype": "Percent",
            "width": 150
        },
        {
            "fieldname": "model",
            "label": _("Model"),
            "fieldtype": "Data",
            "width": 120
        },
        {
            "fieldname": "forecast_date",
            "label": _("Forecast Date"),
            "fieldtype": "Date",
            "width": 150
        }
    ]

def get_data(filters):
    """Get report data"""
    try:
        # Get accuracy data with simplified query
        data = get_accuracy_data_simplified(filters)
        return data
    except Exception as e:
        frappe.log_error(f"Report Data Error: {str(e)}")
        return []

def get_accuracy_data_simplified(filters):
    """Get accuracy data with correct field names"""
    conditions = ["1=1"]  # Remove docstatus filter since records are in draft
    
    if filters:
        if filters.get("company"):
            conditions.append(f"aff.company = '{filters.get('company')}'")
        
        if filters.get("from_date"):
            conditions.append(f"aff.forecast_start_date >= '{filters.get('from_date')}'")
            
        if filters.get("to_date"):
            conditions.append(f"aff.forecast_start_date <= '{filters.get('to_date')}'")  # Use start_date instead of end_date
            
        if filters.get("model_type") and filters.get("model_type") != "All":
            conditions.append(f"aff.prediction_model = '{filters.get('model_type')}'")
    
    where_clause = " AND ".join(conditions) if conditions else "1=1"
    
    # Query joining both tables with correct field names
    query = """
        SELECT 
            aff.name as forecast_id,
            aff.account,
            aff.account_name,
            aff.predicted_amount,
            COALESCE(afa.actual_value, 0) as actual_amount,
            COALESCE(afa.accuracy_percentage, 0) as accuracy_percentage,
            aff.confidence_score,
            aff.prediction_model as model_type,
            aff.forecast_start_date,
            aff.forecast_end_date,
            COALESCE(afa.measurement_date, aff.forecast_start_date) as measurement_date,
            COALESCE(afa.performance_grade, 'Pending') as performance_grade
        FROM `tabAI Financial Forecast` aff
        LEFT JOIN `tabAI Forecast Accuracy` afa ON aff.name = afa.forecast_reference
        WHERE {}
        ORDER BY aff.forecast_start_date DESC
        LIMIT 100
    """.format(where_clause)
    
    return frappe.db.sql(query, as_dict=True)

@frappe.whitelist()
def generate_forecast_accuracy_report(company=None, period_days=90, model_type=None):
    """Generate comprehensive forecast accuracy analysis report"""
    
    try:
        # Get accuracy data
        accuracy_data = get_accuracy_data(company, period_days, model_type)
        
        # Calculate accuracy metrics
        accuracy_metrics = calculate_accuracy_metrics(accuracy_data)
        
        # Get model performance comparison
        model_performance = get_model_performance_comparison(company, period_days)
        
        # Get accuracy trends
        accuracy_trends = get_accuracy_trends(company, period_days)
        
        # Get confidence vs accuracy analysis
        confidence_analysis = get_confidence_vs_accuracy_analysis(company, period_days)
        
        # Get forecast type performance
        type_performance = get_forecast_type_performance(company, period_days)
        
        # Generate insights and recommendations
        insights = generate_accuracy_insights(accuracy_metrics, model_performance, accuracy_trends)
        recommendations = generate_accuracy_recommendations(accuracy_metrics, model_performance)
        
        report_data = {
            "report_title": "AI Forecast Accuracy Analysis Report",
            "generated_at": frappe.utils.now(),
            "company": company or "All Companies",
            "analysis_period": f"Last {period_days} days",
            "model_filter": model_type or "All Models",
            "summary": {
                "total_forecasts_evaluated": len(accuracy_data),
                "overall_accuracy": accuracy_metrics.get("overall_accuracy", 0),
                "average_confidence": accuracy_metrics.get("average_confidence", 0),
                "best_performing_model": model_performance[0]["model"] if model_performance else "N/A",
                "accuracy_trend": accuracy_trends.get("trend_direction", "Stable")
            },
            "accuracy_metrics": accuracy_metrics,
            "model_performance": model_performance,
            "accuracy_trends": accuracy_trends,
            "confidence_analysis": confidence_analysis,
            "type_performance": type_performance,
            "insights": insights,
            "recommendations": recommendations
        }
        
        return {
            "success": True,
            "data": report_data
        }
        
    except Exception as e:
        frappe.log_error(f"Forecast Accuracy Report Error: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }

def get_accuracy_data(company=None, period_days=90, model_type=None):
    """Get accuracy data with error handling for missing tables"""
    
    try:
        # Check if accuracy assessment table exists
        table_exists = frappe.db.sql("""
            SELECT COUNT(*) as count 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() 
            AND table_name = 'tabAI Accuracy Assessment'
        """, as_dict=True)
        
        if not table_exists or table_exists[0].count == 0:
            frappe.log_error("AI Accuracy Assessment table not found, using simplified approach")
            return get_accuracy_data_simplified({"company": company, "period_days": period_days})
        
        # Original complex query with proper field names
        conditions = []
        
        if company:
            conditions.append(f"aff.company = '{company}'")
        
        if period_days:
            date_condition = f"aff.forecast_start_date >= CURDATE() - INTERVAL {period_days} DAY"
            conditions.append(date_condition)
        
        if model_type and model_type != "all":
            conditions.append(f"aff.prediction_model = '{model_type}'")
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        
        # Use forecast_reference instead of forecast_id and correct field names
        accuracy_query = """
        SELECT 
            afa.name,
            afa.forecast_reference,
            afa.predicted_value,
            afa.actual_value,
            afa.accuracy_percentage,
            afa.measurement_date,
            afa.model_used,
            afa.absolute_error,
            afa.percentage_error,
            afa.confidence_calibration as confidence_interval_lower,
            afa.model_reliability_score as confidence_interval_upper,
            afa.model_used,
            aff.prediction_model,
            aff.confidence_score,
            aff.predicted_amount,
            aff.upper_bound,
            aff.lower_bound,
            aff.account,
            aff.account_name,
            aff.company,
            aff.forecast_start_date,
            aff.forecast_end_date
        FROM `tabAI Forecast Accuracy` afa
        JOIN `tabAI Financial Forecast` aff ON afa.forecast_reference = aff.name
        WHERE {}
        ORDER BY afa.measurement_date DESC
        """.format(where_clause)
        
        return frappe.db.sql(accuracy_query, as_dict=True)
        
    except Exception as e:
        frappe.log_error(f"Accuracy Data Error: {str(e)}")
        # Fallback to simplified approach
        return get_accuracy_data_simplified({"company": company, "period_days": period_days})

def calculate_accuracy_metrics(accuracy_data):
    """Calculate comprehensive accuracy metrics"""
    
    if not accuracy_data:
        return {
            "overall_accuracy": 0,
            "average_confidence": 0,
            "total_forecasts": 0,
            "message": "No accuracy data available for analysis"
        }
    
    # Basic metrics
    total_forecasts = len(accuracy_data)
    accuracy_scores = [d.accuracy_score for d in accuracy_data if d.accuracy_score is not None]
    error_percentages = [abs(d.error_percentage) for d in accuracy_data if d.error_percentage is not None]
    confidence_scores = [d.confidence_at_creation for d in accuracy_data if d.confidence_at_creation is not None]
    
    # Calculate accuracy metrics
    overall_accuracy = statistics.mean(accuracy_scores) if accuracy_scores else 0
    median_accuracy = statistics.median(accuracy_scores) if accuracy_scores else 0
    accuracy_std_dev = statistics.stdev(accuracy_scores) if len(accuracy_scores) > 1 else 0
    
    # Calculate error metrics
    mean_absolute_error = statistics.mean(error_percentages) if error_percentages else 0
    median_absolute_error = statistics.median(error_percentages) if error_percentages else 0
    
    # Calculate confidence metrics
    average_confidence = statistics.mean(confidence_scores) if confidence_scores else 0
    
    # Accuracy distribution
    high_accuracy_count = len([a for a in accuracy_scores if a >= 80])
    medium_accuracy_count = len([a for a in accuracy_scores if 60 <= a < 80])
    low_accuracy_count = len([a for a in accuracy_scores if a < 60])
    
    # Calculate forecast horizon impact
    horizon_analysis = analyze_horizon_impact(accuracy_data)
    
    # Calculate confidence-accuracy correlation
    confidence_accuracy_correlation = calculate_confidence_correlation(accuracy_data)
    
    return {
        "total_forecasts": total_forecasts,
        "overall_accuracy": round(overall_accuracy, 2),
        "median_accuracy": round(median_accuracy, 2),
        "accuracy_std_dev": round(accuracy_std_dev, 2),
        "mean_absolute_error": round(mean_absolute_error, 2),
        "median_absolute_error": round(median_absolute_error, 2),
        "average_confidence": round(average_confidence, 2),
        "accuracy_distribution": {
            "high_accuracy": high_accuracy_count,
            "medium_accuracy": medium_accuracy_count, 
            "low_accuracy": low_accuracy_count,
            "high_accuracy_percentage": round((high_accuracy_count / total_forecasts) * 100, 1),
            "medium_accuracy_percentage": round((medium_accuracy_count / total_forecasts) * 100, 1),
            "low_accuracy_percentage": round((low_accuracy_count / total_forecasts) * 100, 1)
        },
        "horizon_analysis": horizon_analysis,
        "confidence_accuracy_correlation": confidence_accuracy_correlation
    }

def analyze_horizon_impact(accuracy_data):
    """Analyze how forecast horizon affects accuracy"""
    
    # Group by forecast horizon
    horizon_groups = {
        "1-7 days": [],
        "8-30 days": [],
        "31-90 days": [],
        "90+ days": []
    }
    
    for data in accuracy_data:
        horizon_days = data.forecast_horizon_days or 0
        
        if horizon_days <= 7:
            horizon_groups["1-7 days"].append(data.accuracy_score)
        elif horizon_days <= 30:
            horizon_groups["8-30 days"].append(data.accuracy_score)
        elif horizon_days <= 90:
            horizon_groups["31-90 days"].append(data.accuracy_score)
        else:
            horizon_groups["90+ days"].append(data.accuracy_score)
    
    # Calculate average accuracy for each horizon
    horizon_analysis = {}
    for horizon, scores in horizon_groups.items():
        if scores:
            horizon_analysis[horizon] = {
                "count": len(scores),
                "average_accuracy": round(statistics.mean(scores), 2),
                "median_accuracy": round(statistics.median(scores), 2)
            }
        else:
            horizon_analysis[horizon] = {
                "count": 0,
                "average_accuracy": 0,
                "median_accuracy": 0
            }
    
    return horizon_analysis

def calculate_confidence_correlation(accuracy_data):
    """Calculate correlation between confidence and accuracy"""
    
    try:
        confidence_scores = []
        accuracy_scores = []
        
        for data in accuracy_data:
            if data.confidence_at_creation is not None and data.accuracy_score is not None:
                confidence_scores.append(data.confidence_at_creation)
                accuracy_scores.append(data.accuracy_score)
        
        if len(confidence_scores) < 2:
            return {"correlation": 0, "interpretation": "Insufficient data"}
        
        # Calculate Pearson correlation coefficient
        n = len(confidence_scores)
        mean_confidence = statistics.mean(confidence_scores)
        mean_accuracy = statistics.mean(accuracy_scores)
        
        numerator = sum((c - mean_confidence) * (a - mean_accuracy) for c, a in zip(confidence_scores, accuracy_scores))
        denominator_c = sum((c - mean_confidence) ** 2 for c in confidence_scores)
        denominator_a = sum((a - mean_accuracy) ** 2 for a in accuracy_scores)
        
        if denominator_c == 0 or denominator_a == 0:
            correlation = 0
        else:
            correlation = numerator / (denominator_c * denominator_a) ** 0.5
        
        # Interpret correlation
        if abs(correlation) >= 0.7:
            interpretation = "Strong correlation"
        elif abs(correlation) >= 0.3:
            interpretation = "Moderate correlation"
        else:
            interpretation = "Weak correlation"
        
        return {
            "correlation": round(correlation, 3),
            "interpretation": interpretation,
            "sample_size": n
        }
        
    except Exception as e:
        return {"correlation": 0, "interpretation": "Calculation error", "error": str(e)}

def get_model_performance_comparison(company=None, period_days=90):
    """Compare performance across different prediction models"""
    
    # Build WHERE conditions directly without parameterization for .format()
    conditions = [f"afa.measurement_date >= DATE_SUB(NOW(), INTERVAL {period_days} DAY)"]
    
    if company:
        conditions.append(f"aff.company = '{company}'")
    
    where_clause = " AND ".join(conditions)
    
    model_performance = frappe.db.sql(f"""
        SELECT 
            aff.prediction_model as model,
            COUNT(*) as forecast_count,
            AVG(afa.accuracy_percentage) as avg_accuracy,
            STDDEV(afa.accuracy_percentage) as accuracy_std_dev,
            AVG(ABS(afa.percentage_error)) as avg_error,
            AVG(aff.confidence_score) as avg_confidence,
            MIN(afa.accuracy_percentage) as min_accuracy,
            MAX(afa.accuracy_percentage) as max_accuracy,
            COUNT(CASE WHEN afa.accuracy_percentage >= 80 THEN 1 END) as high_accuracy_count
        FROM `tabAI Forecast Accuracy` afa
        JOIN `tabAI Financial Forecast` aff ON afa.forecast_reference = aff.name
        WHERE {where_clause}
        AND aff.prediction_model IS NOT NULL
        AND aff.prediction_model != ''
        GROUP BY aff.prediction_model
        ORDER BY avg_accuracy DESC, forecast_count DESC
    """, as_dict=True)
    
    # Calculate performance scores for ranking
    for model in model_performance:
        # Performance score = weighted average of accuracy, consistency, and volume
        accuracy_weight = 0.5
        consistency_weight = 0.3  # Lower std dev = higher consistency
        volume_weight = 0.2
        
        accuracy_score = (model.avg_accuracy or 0) / 100
        consistency_score = 1 - min((model.accuracy_std_dev or 0) / 100, 1)  # Normalize std dev
        volume_score = min((model.forecast_count or 0) / 100, 1)  # Normalize volume
        
        model.performance_score = round(
            (accuracy_score * accuracy_weight) + 
            (consistency_score * consistency_weight) + 
            (volume_score * volume_weight), 3
        )
        
        # Round other values
        model.avg_accuracy = round(model.avg_accuracy or 0, 2)
        model.accuracy_std_dev = round(model.accuracy_std_dev or 0, 2)
        model.avg_error = round(model.avg_error or 0, 2)
        model.avg_confidence = round(model.avg_confidence or 0, 2)
        model.high_accuracy_percentage = round((model.high_accuracy_count / model.forecast_count) * 100, 1)
    
    return sorted(model_performance, key=lambda x: x.performance_score, reverse=True)

def get_accuracy_trends(company=None, period_days=90):
    """Analyze accuracy trends over time"""
    
    # Build WHERE conditions directly without parameterization
    conditions = [f"afa.measurement_date >= DATE_SUB(NOW(), INTERVAL {period_days} DAY)"]
    
    if company:
        conditions.append(f"aff.company = '{company}'")
    
    where_clause = " AND ".join(conditions)
    
    # Get weekly accuracy trends
    weekly_trends = frappe.db.sql(f"""
        SELECT 
            YEARWEEK(afa.measurement_date) as year_week,
            DATE(DATE_SUB(afa.measurement_date, INTERVAL WEEKDAY(afa.measurement_date) DAY)) as week_start,
            AVG(afa.accuracy_percentage) as avg_accuracy,
            COUNT(*) as forecast_count,
            AVG(aff.confidence_score) as avg_confidence
        FROM `tabAI Forecast Accuracy` afa
        JOIN `tabAI Financial Forecast` aff ON afa.forecast_reference = aff.name
        WHERE {where_clause}
        GROUP BY YEARWEEK(afa.measurement_date)
        ORDER BY year_week
    """, as_dict=True)
    
    # Calculate trend direction
    if len(weekly_trends) >= 3:
        recent_accuracy = statistics.mean([w.avg_accuracy for w in weekly_trends[-3:]])
        earlier_accuracy = statistics.mean([w.avg_accuracy for w in weekly_trends[:3]])
        
        if recent_accuracy > earlier_accuracy + 2:
            trend_direction = "Improving"
        elif recent_accuracy < earlier_accuracy - 2:
            trend_direction = "Declining"
        else:
            trend_direction = "Stable"
        
        trend_change = round(recent_accuracy - earlier_accuracy, 2)
    else:
        trend_direction = "Insufficient Data"
        trend_change = 0
    
    # Round values for display
    for trend in weekly_trends:
        trend.avg_accuracy = round(trend.avg_accuracy or 0, 2)
        trend.avg_confidence = round(trend.avg_confidence or 0, 2)
    
    return {
        "weekly_trends": weekly_trends,
        "trend_direction": trend_direction,
        "trend_change": trend_change,
        "analysis_period_weeks": len(weekly_trends)
    }

def get_confidence_vs_accuracy_analysis(company=None, period_days=90):
    """Analyze relationship between confidence scores and actual accuracy"""
    
    # Build WHERE conditions directly without parameterization
    conditions = [f"afa.measurement_date >= DATE_SUB(NOW(), INTERVAL {period_days} DAY)"]
    
    if company:
        conditions.append(f"aff.company = '{company}'")
    
    where_clause = " AND ".join(conditions)
    
    # Group forecasts by confidence ranges
    confidence_analysis = frappe.db.sql(f"""
        SELECT 
            CASE 
                WHEN aff.confidence_score >= 80 THEN 'High (80-100%)'
                WHEN aff.confidence_score >= 60 THEN 'Medium (60-79%)'
                WHEN aff.confidence_score >= 40 THEN 'Low (40-59%)'
                ELSE 'Very Low (<40%)'
            END as confidence_range,
            COUNT(*) as forecast_count,
            AVG(afa.accuracy_percentage) as avg_accuracy,
            AVG(aff.confidence_score) as avg_confidence,
            AVG(ABS(afa.percentage_error)) as avg_error,
            COUNT(CASE WHEN afa.accuracy_percentage >= 80 THEN 1 END) as high_accuracy_count
        FROM `tabAI Forecast Accuracy` afa
        JOIN `tabAI Financial Forecast` aff ON afa.forecast_reference = aff.name
        WHERE {where_clause}
        AND aff.confidence_score IS NOT NULL
        AND afa.accuracy_percentage IS NOT NULL
        GROUP BY 
            CASE 
                WHEN aff.confidence_score >= 80 THEN 'High (80-100%)'
                WHEN aff.confidence_score >= 60 THEN 'Medium (60-79%)'
                WHEN aff.confidence_score >= 40 THEN 'Low (40-59%)'
                ELSE 'Very Low (<40%)'
            END
        ORDER BY avg_confidence DESC
    """, as_dict=True)
    
    # Calculate calibration metrics
    for analysis in confidence_analysis:
        analysis.avg_accuracy = round(analysis.avg_accuracy or 0, 2)
        analysis.avg_confidence = round(analysis.avg_confidence or 0, 2)
        analysis.avg_error = round(analysis.avg_error or 0, 2)
        analysis.high_accuracy_percentage = round((analysis.high_accuracy_count / analysis.forecast_count) * 100, 1)
        
        # Calibration score - how well confidence predicts accuracy
        confidence_accuracy_diff = abs(analysis.avg_confidence - analysis.avg_accuracy)
        analysis.calibration_score = max(0, 100 - confidence_accuracy_diff)
    
    return {
        "confidence_ranges": confidence_analysis,
        "overall_calibration": round(statistics.mean([a.calibration_score for a in confidence_analysis]), 2) if confidence_analysis else 0
    }

def get_forecast_type_performance(company=None, period_days=90):
    """Analyze performance by forecast type"""
    
    # Build WHERE conditions directly without parameterization
    conditions = [f"aa.measurement_date >= DATE_SUB(NOW(), INTERVAL {period_days} DAY)"]
    
    if company:
        conditions.append(f"aff.company = '{company}'")
    
    where_clause = " AND ".join(conditions)
    
    type_performance = frappe.db.sql(f"""
        SELECT 
            afa.forecast_type,
            COUNT(*) as forecast_count,
            AVG(afa.accuracy_percentage) as avg_accuracy,
            STDDEV(afa.accuracy_percentage) as accuracy_std_dev,
            AVG(ABS(afa.percentage_error)) as avg_error,
            AVG(aff.confidence_score) as avg_confidence,
            COUNT(CASE WHEN afa.accuracy_percentage >= 80 THEN 1 END) as high_accuracy_count,
            AVG(DATEDIFF(afa.measurement_date, aff.forecast_start_date)) as avg_horizon_days
        FROM `tabAI Forecast Accuracy` afa
        JOIN `tabAI Financial Forecast` aff ON afa.forecast_reference = aff.name
        WHERE {where_clause}
        GROUP BY afa.forecast_type
        ORDER BY avg_accuracy DESC
    """, as_dict=True)
    
    # Calculate performance metrics
    for perf in type_performance:
        perf.avg_accuracy = round(perf.avg_accuracy or 0, 2)
        perf.accuracy_std_dev = round(perf.accuracy_std_dev or 0, 2)
        perf.avg_error = round(perf.avg_error or 0, 2)
        perf.avg_confidence = round(perf.avg_confidence or 0, 2)
        perf.avg_horizon_days = round(perf.avg_horizon_days or 0, 1)
        perf.high_accuracy_percentage = round((perf.high_accuracy_count / perf.forecast_count) * 100, 1)
        
        # Performance grade
        if perf.avg_accuracy >= 85:
            perf.performance_grade = "Excellent"
        elif perf.avg_accuracy >= 75:
            perf.performance_grade = "Good"
        elif perf.avg_accuracy >= 65:
            perf.performance_grade = "Fair"
        else:
            perf.performance_grade = "Needs Improvement"
    
    return type_performance

def generate_accuracy_insights(accuracy_metrics, model_performance, accuracy_trends):
    """Generate insights from accuracy analysis"""
    
    insights = []
    
    # Overall accuracy insight
    overall_accuracy = accuracy_metrics.get("overall_accuracy", 0)
    if overall_accuracy >= 85:
        insights.append({
            "category": "Overall Performance",
            "type": "positive",
            "title": "Excellent Forecast Accuracy",
            "description": f"Overall accuracy of {overall_accuracy}% indicates highly reliable forecasting",
            "impact": "High confidence in financial planning decisions"
        })
    elif overall_accuracy >= 75:
        insights.append({
            "category": "Overall Performance",
            "type": "neutral",
            "title": "Good Forecast Accuracy",
            "description": f"Overall accuracy of {overall_accuracy}% is within acceptable range",
            "impact": "Generally reliable for strategic planning"
        })
    else:
        insights.append({
            "category": "Overall Performance",
            "type": "negative",
            "title": "Below-Target Accuracy",
            "description": f"Overall accuracy of {overall_accuracy}% needs improvement",
            "impact": "May affect quality of financial decisions"
        })
    
    # Model performance insight
    if model_performance:
        best_model = model_performance[0]
        worst_model = model_performance[-1]
        
        if len(model_performance) > 1 and best_model.avg_accuracy - worst_model.avg_accuracy > 10:
            accuracy_difference = best_model.avg_accuracy - worst_model.avg_accuracy
            insights.append({
                "category": "Model Performance",
                "type": "actionable",
                "title": "Significant Model Performance Variation",
                "description": f"{best_model.model} ({best_model.avg_accuracy}%) outperforms {worst_model.model} ({worst_model.avg_accuracy}%) by {accuracy_difference:.1f} percentage points",
                "impact": "Consider standardizing on best-performing model"
            })
    
    # Trend insight
    trend_direction = accuracy_trends.get("trend_direction", "Stable")
    trend_change = accuracy_trends.get("trend_change", 0)
    
    if trend_direction == "Improving":
        insights.append({
            "category": "Trends",
            "type": "positive",
            "title": "Improving Accuracy Trend",
            "description": f"Accuracy has improved by {trend_change:.1f} percentage points recently",
            "impact": "Forecasting models are learning and adapting well"
        })
    elif trend_direction == "Declining":
        insights.append({
            "category": "Trends",
            "type": "negative", 
            "title": "Declining Accuracy Trend",
            "description": f"Accuracy has declined by {abs(trend_change):.1f} percentage points recently",
            "impact": "Requires investigation and model retraining"
        })
    
    # Confidence-accuracy correlation insight
    correlation = accuracy_metrics.get("confidence_accuracy_correlation", {}).get("correlation", 0)
    if abs(correlation) < 0.3:
        insights.append({
            "category": "Model Calibration",
            "type": "negative",
            "title": "Poor Confidence Calibration",
            "description": "Confidence scores don't correlate well with actual accuracy",
            "impact": "Difficulty in assessing forecast reliability"
        })
    elif correlation > 0.7:
        insights.append({
            "category": "Model Calibration", 
            "type": "positive",
            "title": "Well-Calibrated Confidence Scores",
            "description": "Confidence scores accurately reflect forecast reliability",
            "impact": "Can trust confidence scores for decision making"
        })
    
    return insights

def generate_accuracy_recommendations(accuracy_metrics, model_performance):
    """Generate actionable recommendations for improving accuracy"""
    
    recommendations = []
    
    overall_accuracy = accuracy_metrics.get("overall_accuracy", 0)
    
    # Overall accuracy recommendations
    if overall_accuracy < 75:
        recommendations.append({
            "priority": "High",
            "category": "Model Improvement",
            "title": "Improve Overall Forecast Accuracy",
            "description": f"Current accuracy of {overall_accuracy}% is below target (75%+)",
            "actions": [
                "Review and clean historical training data",
                "Implement ensemble modeling techniques",
                "Increase model training frequency",
                "Add more relevant feature variables",
                "Consider advanced ML algorithms (LSTM, Prophet)"
            ],
            "expected_impact": "10-15% accuracy improvement",
            "timeline": "2-4 weeks",
            "resources_needed": "Data Science team, additional compute resources"
        })
    
    # Model-specific recommendations
    if model_performance and len(model_performance) > 1:
        best_model = model_performance[0]
        underperforming_models = [m for m in model_performance if m.avg_accuracy < best_model.avg_accuracy - 5]
        
        if underperforming_models:
            recommendations.append({
                "priority": "Medium",
                "category": "Model Optimization",
                "title": "Optimize Underperforming Models",
                "description": f"Replace or retrain {len(underperforming_models)} underperforming models",
                "actions": [
                    f"Adopt {best_model.model} approach for underperforming models",
                    "Analyze why certain models perform better",
                    "Implement model selection based on data characteristics",
                    "Regular model performance monitoring"
                ],
                "expected_impact": "5-10% accuracy improvement",
                "timeline": "3-6 weeks",
                "resources_needed": "ML Engineering team"
            })
    
    # Data quality recommendations
    data_quality_score = accuracy_metrics.get("confidence_accuracy_correlation", {}).get("correlation", 0)
    if abs(data_quality_score) < 0.3:
        recommendations.append({
            "priority": "Medium",
            "category": "Data Quality",
            "title": "Improve Data Quality and Model Calibration",
            "description": "Poor correlation between confidence and accuracy indicates data or calibration issues",
            "actions": [
                "Implement automated data quality checks",
                "Review feature engineering process",
                "Calibrate model confidence scores",
                "Add data validation rules",
                "Monitor for data drift"
            ],
            "expected_impact": "Better reliability assessment",
            "timeline": "2-3 weeks",
            "resources_needed": "Data Engineering team"
        })
    
    # Horizon-specific recommendations
    horizon_analysis = accuracy_metrics.get("horizon_analysis", {})
    long_term_accuracy = horizon_analysis.get("90+ days", {}).get("average_accuracy", 0)
    short_term_accuracy = horizon_analysis.get("1-7 days", {}).get("average_accuracy", 0)
    
    if long_term_accuracy > 0 and short_term_accuracy - long_term_accuracy > 15:
        recommendations.append({
            "priority": "Low",
            "category": "Forecast Horizon",
            "title": "Improve Long-term Forecast Accuracy",
            "description": f"Long-term forecasts ({long_term_accuracy}%) significantly less accurate than short-term ({short_term_accuracy}%)",
            "actions": [
                "Implement separate models for different time horizons",
                "Add seasonal and trend components for long-term forecasts",
                "Regular model retraining for long-term predictions",
                "Include external economic indicators"
            ],
            "expected_impact": "Improved long-term planning capability",
            "timeline": "4-8 weeks",
            "resources_needed": "Advanced analytics capabilities"
        })
    
    return recommendations

@frappe.whitelist()
def get_accuracy_details_for_forecast(forecast_id):
    """Get detailed accuracy information for a specific forecast"""
    
    try:
        accuracy_details = frappe.db.sql("""
            SELECT 
                afa.*,
                aff.prediction_model,
                aff.confidence_score,
                aff.predicted_amount,
                aff.upper_bound,
                aff.lower_bound,
                aff.account_name,
                aff.forecast_type
            FROM `tabAI Forecast Accuracy` afa
            JOIN `tabAI Financial Forecast` aff ON afa.forecast_reference = aff.name
            WHERE afa.forecast_reference = %s
        """, (forecast_id,), as_dict=True)
        
        if accuracy_details:
            detail = accuracy_details[0]
            
            # Calculate additional metrics
            if detail.predicted_value and detail.actual_value:
                absolute_error = abs(detail.predicted_value - detail.actual_value)
                relative_error = (absolute_error / abs(detail.actual_value)) * 100 if detail.actual_value != 0 else 0
                
                detail.absolute_error = absolute_error
                detail.relative_error = round(relative_error, 2)
                
                # Determine accuracy category
                if detail.accuracy_score >= 90:
                    detail.accuracy_category = "Excellent"
                elif detail.accuracy_score >= 80:
                    detail.accuracy_category = "Good"
                elif detail.accuracy_score >= 70:
                    detail.accuracy_category = "Fair"
                else:
                    detail.accuracy_category = "Poor"
            
            return {
                "success": True,
                "data": detail
            }
        else:
            return {
                "success": False,
                "message": "No accuracy data found for this forecast"
            }
            
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

@frappe.whitelist()
def export_accuracy_report(company=None, period_days=90, format="excel"):
    """Export accuracy report in specified format"""
    
    report_data = generate_forecast_accuracy_report(company, period_days)
    
    if not report_data["success"]:
        return report_data
    
    if format == "excel":
        return export_accuracy_to_excel(report_data["data"])
    elif format == "pdf":
        return export_accuracy_to_pdf(report_data["data"])
    else:
        return report_data

def export_accuracy_to_excel(data):
    """Export accuracy report to Excel"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import LineChart, Reference
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ["Forecast Accuracy Analysis Report", ""],
            ["Generated", data["generated_at"]],
            ["Company", data["company"]],
            ["Analysis Period", data["analysis_period"]],
            ["", ""],
            ["Key Metrics", ""],
            ["Total Forecasts Evaluated", data["summary"]["total_forecasts_evaluated"]],
            ["Overall Accuracy", f"{data['summary']['overall_accuracy']}%"],
            ["Average Confidence", f"{data['summary']['average_confidence']}%"],
            ["Best Performing Model", data["summary"]["best_performing_model"]],
            ["Accuracy Trend", data["summary"]["accuracy_trend"]]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                if col_num == 1 and value and "Report" not in str(value):
                    cell.font = Font(bold=True)
        
        # Model Performance sheet
        ws_models = wb.create_sheet("Model Performance")
        
        model_headers = ["Model", "Forecasts", "Avg Accuracy", "Std Dev", "Avg Error", "Avg Confidence", "Performance Score"]
        for col_num, header in enumerate(model_headers, 1):
            cell = ws_models.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, model in enumerate(data["model_performance"], 2):
            ws_models.cell(row=row_num, column=1, value=model["model"])
            ws_models.cell(row=row_num, column=2, value=model["forecast_count"])
            ws_models.cell(row=row_num, column=3, value=f"{model['avg_accuracy']}%")
            ws_models.cell(row=row_num, column=4, value=f"{model['accuracy_std_dev']}%")
            ws_models.cell(row=row_num, column=5, value=f"{model['avg_error']}%")
            ws_models.cell(row=row_num, column=6, value=f"{model['avg_confidence']}%")
            ws_models.cell(row=row_num, column=7, value=model["performance_score"])
        
        # Accuracy Trends sheet
        if data["accuracy_trends"]["weekly_trends"]:
            ws_trends = wb.create_sheet("Accuracy Trends")
            
            trend_headers = ["Week Start", "Avg Accuracy", "Forecast Count", "Avg Confidence"]
            for col_num, header in enumerate(trend_headers, 1):
                cell = ws_trends.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row_num, trend in enumerate(data["accuracy_trends"]["weekly_trends"], 2):
                ws_trends.cell(row=row_num, column=1, value=trend["week_start"])
                ws_trends.cell(row=row_num, column=2, value=f"{trend['avg_accuracy']}%")
                ws_trends.cell(row=row_num, column=3, value=trend["forecast_count"])
                ws_trends.cell(row=row_num, column=4, value=f"{trend['avg_confidence']}%")
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            "success": True,
            "content": output.getvalue(),
            "filename": f"forecast_accuracy_report_{frappe.utils.today()}.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
    except Exception as e:
        frappe.log_error(f"Excel export error: {str(e)}")
        return {"success": False, "error": str(e)}

def export_accuracy_to_pdf(data):
    """Export accuracy report to PDF"""
    
    html_content = generate_accuracy_html_report(data)
    
    try:
        pdf_content = frappe.utils.get_pdf(html_content)
        return {
            "success": True,
            "content": pdf_content,
            "filename": f"forecast_accuracy_report_{frappe.utils.today()}.pdf",
            "content_type": "application/pdf"
        }
    except Exception as e:
        frappe.log_error(f"PDF export error: {str(e)}")
        return {"success": False, "error": str(e)}

def generate_accuracy_html_report(data):
    """Generate HTML version of accuracy report"""
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Forecast Accuracy Analysis Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
            .header {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
            .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }}
            .metric-card {{ border: 1px solid #ddd; padding: 15px; border-radius: 5px; text-align: center; }}
            .excellent {{ background-color: #d4edda; }}
            .good {{ background-color: #d1ecf1; }}
            .fair {{ background-color: #fff3cd; }}
            .poor {{ background-color: #f8d7da; }}
            table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            .insight {{ margin: 10px 0; padding: 15px; border-radius: 5px; }}
            .insight.positive {{ background-color: #d4edda; border-left: 5px solid #28a745; }}
            .insight.negative {{ background-color: #f8d7da; border-left: 5px solid #dc3545; }}
            .insight.neutral {{ background-color: #d1ecf1; border-left: 5px solid #17a2b8; }}
            .insight.actionable {{ background-color: #fff3cd; border-left: 5px solid #ffc107; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Forecast Accuracy Analysis Report</h1>
            <p><strong>Company:</strong> {data['company']}</p>
            <p><strong>Generated:</strong> {data['generated_at']}</p>
            <p><strong>Analysis Period:</strong> {data['analysis_period']}</p>
        </div>
        
        <div class="section">
            <h2>Executive Summary</h2>
            <div class="metric-grid">
                <div class="metric-card">
                    <h3>{data['summary']['total_forecasts_evaluated']}</h3>
                    <p>Forecasts Evaluated</p>
                </div>
                <div class="metric-card">
                    <h3>{data['summary']['overall_accuracy']}%</h3>
                    <p>Overall Accuracy</p>
                </div>
                <div class="metric-card">
                    <h3>{data['summary']['average_confidence']}%</h3>
                    <p>Average Confidence</p>
                </div>
                <div class="metric-card">
                    <h3>{data['summary']['accuracy_trend']}</h3>
                    <p>Accuracy Trend</p>
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2>Model Performance Comparison</h2>
            <table>
                <thead>
                    <tr>
                        <th>Model</th>
                        <th>Forecasts</th>
                        <th>Avg Accuracy</th>
                        <th>Std Deviation</th>
                        <th>Performance Score</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for model in data["model_performance"]:
        html += f"""
                    <tr>
                        <td>{model['model']}</td>
                        <td>{model['forecast_count']}</td>
                        <td>{model['avg_accuracy']}%</td>
                        <td>{model['accuracy_std_dev']}%</td>
                        <td>{model['performance_score']}</td>
                    </tr>
        """
    
    html += """
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2>Key Insights</h2>
    """
    
    for insight in data["insights"]:
        html += f"""
            <div class="insight {insight['type']}">
                <h4>{insight['title']}</h4>
                <p>{insight['description']}</p>
                <small><strong>Impact:</strong> {insight['impact']}</small>
            </div>
        """
    
    html += """
        </div>
        
        <div class="section">
            <h2>Recommendations</h2>
    """
    
    for rec in data["recommendations"]:
        html += f"""
            <div class="recommendation" style="margin-bottom: 20px; padding: 15px; border: 1px solid #ddd; border-radius: 5px;">
                <h4>{rec['title']} ({rec['priority']} Priority)</h4>
                <p>{rec['description']}</p>
                <p><strong>Expected Impact:</strong> {rec['expected_impact']}</p>
                <p><strong>Timeline:</strong> {rec['timeline']}</p>
                <p><strong>Resources Needed:</strong> {rec['resources_needed']}</p>
            </div>
        """
    
    html += """
        </div>
    </body>
    </html>
    """
    
    return html

@frappe.whitelist()
def get_model_analysis(company=None, period_days=90):
    """Get detailed model analysis data"""
    try:
        # Use existing function to get model performance
        model_performance = get_model_performance_comparison(company, period_days)
        
        return {
            "success": True,
            "data": {
                "model_performance": model_performance,
                "total_models": len(model_performance),
                "analysis_period": f"Last {period_days} days"
            }
        }
    except Exception as e:
        frappe.log_error(f"Model Analysis Error: {str(e)}")
        return {"success": False, "error": str(e)}

@frappe.whitelist()
def get_detailed_analysis(company=None, period_days=90, model_type=None):
    """Get detailed analysis for specific criteria"""
    try:
        # Get comprehensive report data
        report_data = generate_forecast_accuracy_report(company, period_days, model_type)
        
        if report_data.get("success"):
            return {
                "success": True,
                "data": report_data.get("data", {})
            }
        else:
            return report_data
            
    except Exception as e:
        frappe.log_error(f"Detailed Analysis Error: {str(e)}")
        return {"success": False, "error": str(e)}

@frappe.whitelist()
def get_quick_accuracy_summary(company=None):
    """Get quick accuracy summary for dashboard"""
    try:
        # Get recent accuracy data (last 30 days)
        accuracy_data = get_accuracy_data(company, 30)
        
        if not accuracy_data:
            return {
                "success": True,
                "data": {
                    "total_forecasts": 0,
                    "overall_accuracy": 0,
                    "trend": "No data"
                }
            }
        
        # Calculate basic metrics
        accuracy_values = [d.get("accuracy_percentage", 0) for d in accuracy_data if d.get("accuracy_percentage") is not None]
        
        summary = {
            "total_forecasts": len(accuracy_data),
            "overall_accuracy": statistics.mean(accuracy_values) if accuracy_values else 0,
            "best_accuracy": max(accuracy_values) if accuracy_values else 0,
            "worst_accuracy": min(accuracy_values) if accuracy_values else 0,
            "trend": "Stable"  # Simplified for now
        }
        
        return {"success": True, "data": summary}
        
    except Exception as e:
        frappe.log_error(f"Quick Summary Error: {str(e)}")
        return {"success": False, "error": str(e)}

@frappe.whitelist()
def get_model_rankings(company=None, period_days=90):
    """Get model rankings by performance"""
    try:
        # Get model performance data
        model_performance = get_model_performance_comparison(company, period_days)
        
        # Sort by accuracy
        model_performance.sort(key=lambda x: x.get("accuracy", 0), reverse=True)
        
        # Add rankings
        for i, model in enumerate(model_performance):
            model["rank"] = i + 1
        
        return {
            "success": True,
            "data": {
                "rankings": model_performance,
                "total_models": len(model_performance)
            }
        }
        
    except Exception as e:
        frappe.log_error(f"Model Rankings Error: {str(e)}")
        return {"success": False, "error": str(e)}